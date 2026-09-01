"""Additional edge-case tests (run with `python -m unittest` from the package root).

These lock down behaviors surfaced while reviewing real documents:
Luhn card filtering, SSN-vs-organizational-ID distinction, HTML signature
recovery (the .msg fix), category mapping, keyword detection, name formats,
the full lane-routing table, bucket edges, estimation, and legacy rule packs.
"""
import csv
import json
import os
import re
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from pii_triage.config import DEFAULT_RULEPACK, load_rulepack, Config, load_dotenv
from pii_triage.benchmark import (run_benchmark, _parse_resp, _pick_column,
                                  _pred_from_lane, _ID_HINTS, _RESP_HINTS)
from pii_triage.detection import (CompiledRules, detect, detect_names,
                                  detect_addresses, detect_ssn,
                                  luhn_valid, card_valid, value_signal)
from pii_triage.azure_clients import _sample_for_llm
from pii_triage.enrich import apply_ocr, apply_llm
from pii_triage.extractors import _strip_html
from pii_triage.routing import (bucket_of, choose_lane, classify_ambiguity,
                                estimate_entities, FileRecord)

RULES = CompiledRules.from_pack(DEFAULT_RULEPACK, use_ner=False)


class TestCardLuhn(unittest.TestCase):
    """A card is only a card if it passes the Luhn checksum -- guards against
    flagging arbitrary digit runs (invoice numbers, IDs) as payment cards."""

    def test_valid_card_detected(self):
        counts, labels, _ = detect("payment 4111 1111 1111 1111 today", RULES)
        self.assertEqual(counts["CARD"], 1)
        self.assertIn("Payment Card", labels)

    def test_non_luhn_number_not_a_card(self):
        counts, _, _ = detect("ref 1234 5678 9012 3456 end", RULES)
        self.assertEqual(counts["CARD"], 0)

    def test_short_number_not_a_card(self):
        # fewer than 13 digits cannot be a card number
        counts, _, _ = detect("invoice 12345678", RULES)
        self.assertEqual(counts["CARD"], 0)


class TestGovIdVsOrgId(unittest.TestCase):
    """Organizational tax IDs / charity numbers must not be read as an SSN."""

    def test_ssn_detected(self):
        counts, _, _ = detect("SSN 123-45-6789", RULES)
        self.assertEqual(counts["SSN"], 1)

    def test_ein_not_ssn(self):
        # EIN is 2-7 digits, not SSN's 3-2-4 shape
        counts, _, _ = detect("EIN 36-2193610 on file", RULES)
        self.assertEqual(counts["SSN"], 0)

    def test_charity_number_not_ssn(self):
        counts, _, _ = detect("Registered charity 839160801 RR0001", RULES)
        self.assertEqual(counts["SSN"], 0)


class TestHtmlSignatureRecovery(unittest.TestCase):
    """The .msg fix: emails/phones present only as HTML links in a signature
    must be recoverable. Mirrors x_msg's strip + mailto/tel extraction."""

    def _msg_like_text(self, html):
        links = " ".join(re.findall(r"(?:mailto:|tel:)([^\"'>\s]+)", html, re.I))
        return _strip_html(html) + " " + links

    def test_email_link_recovered(self):
        text = self._msg_like_text('<a href="mailto:jane@cng-inc.com">Email Jane</a>')
        counts, _, _ = detect(text, RULES)
        self.assertEqual(counts["EMAIL"], 1)

    def test_tel_link_recovered(self):
        text = self._msg_like_text('<a href="tel:9146969024">Call</a>')
        counts, _, _ = detect(text, RULES)
        self.assertEqual(counts["PHONE"], 1)

    def test_visible_html_text_survives(self):
        _, labels, _ = detect(_strip_html("<p>diagnosis noted</p>"), RULES)
        self.assertIn("Health", labels)


class TestCategoryMapping(unittest.TestCase):
    def test_multi_category(self):
        text = "John Smith SSN 123-45-6789 email j@k.com diagnosis employee id 7"
        _, _, cats = detect(text, RULES)
        self.assertIn("Government-Issued Identification", cats)
        self.assertIn("Contact Information", cats)
        self.assertIn("Health-Related Information", cats)
        self.assertIn("Work-Related Information", cats)

    def test_name_has_no_category(self):
        # NAME is the data-subject identity, not one of the protocol PI-Type categories
        _, labels, cats = detect("Mr. Robert Jones", RULES)
        self.assertIn("Name", labels)
        self.assertEqual(cats, [])


class TestKeywordCategories(unittest.TestCase):
    def test_health(self):
        _, labels, _ = detect("patient diagnosis and treatment plan", RULES)
        self.assertIn("Health", labels)

    def test_family(self):
        _, labels, _ = detect("mother's maiden name required", RULES)
        self.assertIn("Family", labels)

    def test_demographic(self):
        _, labels, _ = detect("religion and ethnicity recorded", RULES)
        self.assertIn("Demographic", labels)

    def test_work(self):
        _, labels, _ = detect("employee id 4821 and salary details", RULES)
        self.assertIn("Work-Related", labels)

    def test_credentials(self):
        _, labels, _ = detect("username and password reset", RULES)
        self.assertIn("Credentials", labels)


class TestNameFormats(unittest.TestCase):
    def test_titled_name(self):
        self.assertIn("robert jones", detect_names("Mr. Robert Jones", use_ner=False))

    def test_labeled_name(self):
        self.assertIn("jane doe", detect_names("Name: Jane Doe", use_ner=False))

    def test_bare_name_not_guessed(self):
        # the whole point: no hardcoded first-name list; bare pairs aren't guessed
        self.assertEqual(detect_names("spoke with Andrew Wallach", use_ner=False), set())

    def test_org_phrase_not_a_person(self):
        self.assertEqual(detect_names("Forest Legacy Foundation donation", use_ner=False), set())


class TestAddressPrecision(unittest.TestCase):
    def test_street_detected(self):
        self.assertEqual(detect_addresses("123 Main Street, Springfield"), 1)

    def test_city_state_zip_detected(self):
        self.assertEqual(detect_addresses("Purchase, NY 10577"), 1)

    def test_bare_five_digits_not_an_address(self):
        # ID-number fragments must not be read as a ZIP/address
        self.assertEqual(detect_addresses("CRA #72740 9880 RR0001 and 10774 6174"), 0)


class TestOrgIdNotPersonal(unittest.TestCase):
    def test_ein_not_personal_taxid(self):
        counts, _, _ = detect("EIN 36-2193610 on file", RULES)
        self.assertEqual(counts["TIN"], 0)


class TestCardReference(unittest.TestCase):
    def test_amex_reference(self):
        _, labels, _ = detect("I'll put it on your AMEX", RULES)
        self.assertIn("Card/Transaction", labels)

    def test_credit_card_phrase(self):
        _, labels, cats = detect("charged to your credit card", RULES)
        self.assertIn("Card/Transaction", labels)
        self.assertIn("Financial Account Information", cats)

    def test_visa_word_not_flagged(self):
        # bare "visa" (immigration context) is intentionally NOT a card keyword
        _, labels, _ = detect("visa application for travel", RULES)
        self.assertNotIn("Card/Transaction", labels)


class TestAmbiguityGate(unittest.TestCase):
    def test_strong_id_is_clear_responsive(self):
        counts, labels, _ = detect("SSN 123-45-6789", RULES)
        self.assertEqual(classify_ambiguity(counts, labels, False), "clear_responsive")

    def test_nothing_is_clear_non_responsive(self):
        counts, labels, _ = detect("internal memo, no personal data", RULES)
        self.assertEqual(classify_ambiguity(counts, labels, False), "clear_non_responsive")

    def test_contact_only_is_ambiguous(self):
        counts, labels, _ = detect("reach me at a@b.com", RULES)
        self.assertEqual(classify_ambiguity(counts, labels, False), "ambiguous")

    def test_card_alone_is_ambiguous_not_strong(self):
        # a card-shaped number must go to the AI (could be an order/transaction id)
        counts, labels, _ = detect("card 4532 0151 1283 0366", RULES)
        self.assertGreater(counts.get("CARD", 0), 0)
        self.assertEqual(classify_ambiguity(counts, labels, False), "ambiguous")

    def test_structured_with_pii_is_ambiguous(self):
        counts, labels, _ = detect("contact a@b.com", RULES)
        self.assertEqual(classify_ambiguity(counts, labels, True), "ambiguous")

    def test_structured_with_identifier_rows_is_ambiguous(self):
        counts, labels, _ = detect("anything", RULES)   # no labels, but rows exist
        self.assertEqual(
            classify_ambiguity(counts, labels, True, structured_rows=5), "ambiguous")

    def test_structured_empty_is_non_responsive(self):
        counts, labels, _ = detect("anything", RULES)
        self.assertEqual(
            classify_ambiguity(counts, labels, True, structured_rows=0),
            "clear_non_responsive")

    def test_structured_with_ssn_is_clear_responsive(self):
        counts, labels, _ = detect("SSN 123-45-6789", RULES)
        self.assertEqual(classify_ambiguity(counts, labels, True), "clear_responsive")


class TestLaneRouting(unittest.TestCase):
    def _rec(self, **kw):
        base = dict(rel_path="a", file_name="a", ext=".x", size_bytes=1)
        base.update(kw)
        return FileRecord(**base)

    def test_standard(self):
        r = self._rec(status="ok", searchable=True, estimated_entities=3,
                      entities_found="Name | Email", value_signal=True)
        self.assertEqual(choose_lane(r), "standard")

    def test_likely_non_responsive(self):
        r = self._rec(status="ok", searchable=True, estimated_entities=0, entities_found="")
        self.assertEqual(choose_lane(r), "likely_non_responsive")

    def test_mention_only_is_not_responsive_without_llm(self):
        # A bare topic mention (no actual value, no LLM call) must NOT flag: this is
        # the over-call we are trimming. estimated_entities is the keyword floor of 1.
        r = self._rec(status="ok", searchable=True, estimated_entities=1,
                      entities_found="Card/Transaction", value_signal=False)
        self.assertEqual(choose_lane(r), "likely_non_responsive")

    def test_name_only_is_not_responsive_without_llm(self):
        # A name on its own is not responsive (protocol); it must not auto-flag.
        r = self._rec(status="ok", searchable=True, estimated_entities=2,
                      entities_found="Name", value_signal=False)
        self.assertEqual(choose_lane(r), "likely_non_responsive")

    def test_value_signal_flags_when_llm_off(self):
        # A real personal value still flags via the rules fallback -- never undercall.
        r = self._rec(status="ok", searchable=True, estimated_entities=1,
                      entities_found="Email", value_signal=True)
        self.assertEqual(choose_lane(r), "standard")

    def test_structured_rows_flag_when_llm_off(self):
        # Identifier-bearing structured rows flag even with no value_signal label.
        r = self._rec(status="ok", searchable=True, estimated_entities=12,
                      is_structured=True, value_signal=False)
        self.assertEqual(choose_lane(r), "standard")

    def test_bde_unstructured(self):
        r = self._rec(status="ok", searchable=True, estimated_entities=80,
                      entities_found="SSN", value_signal=True, is_bde=True,
                      is_structured=False)
        self.assertEqual(choose_lane(r), "bde")

    def test_structured_bde(self):
        r = self._rec(status="ok", searchable=True, estimated_entities=80,
                      entities_found="SSN", is_bde=True, is_structured=True)
        self.assertEqual(choose_lane(r), "structured_bde")

    def test_nonsearchable(self):
        r = self._rec(status="ok", searchable=False, text_extractable="image_only")
        self.assertEqual(choose_lane(r), "nonsearchable_sample")

    def test_container(self):
        self.assertEqual(choose_lane(self._rec(status="container")), "container_expand")

    def test_convert_lane(self):
        r = self._rec(status="no_parser", text_extractable="needs_conversion")
        self.assertEqual(choose_lane(r), "convert_lane")

    def test_needs_parser(self):
        r = self._rec(status="no_parser", text_extractable="unknown")
        self.assertEqual(choose_lane(r), "needs_parser")

    def test_oversize(self):
        self.assertEqual(choose_lane(self._rec(status="skipped_too_large")), "manual_oversize")

    def test_error(self):
        self.assertEqual(choose_lane(self._rec(status="error", detail="boom")), "review_error")


class TestValueSignal(unittest.TestCase):
    """value_signal separates an actual personal value from a bare topic mention --
    the recall-safe responsiveness floor when the LLM is off or its call fails."""

    def _vs(self, text):
        counts, _labels, _cats = detect(text, RULES)
        return value_signal(counts, RULES)

    def test_real_value_is_signal(self):
        self.assertTrue(self._vs("reach me at jane@home.com"))
        self.assertTrue(self._vs("SSN 123-45-6789"))

    def test_bare_mention_is_not_signal(self):
        # generic topic references with no actual personal value -> over-calls trimmed
        self.assertFalse(self._vs("we accept credit cards and debit cards"))
        self.assertFalse(self._vs("please reset your password on first login"))
        self.assertFalse(self._vs("Patient intake form -- Date of Birth: ____"))

    def test_name_alone_is_not_signal(self):
        self.assertFalse(self._vs("Dear Mr. John Smith,"))

    def test_name_with_other_signal_is_signal(self):
        # a name WITH a PI topic (health) is responsive even with no parsed value
        self.assertTrue(self._vs("Patient: Mr. John Smith. Prescription enclosed."))

    def test_nothing_is_not_signal(self):
        self.assertFalse(self._vs("internal logistics memo, no personal data"))


class TestEstimateAndBuckets(unittest.TestCase):
    def test_bucket_edges(self):
        self.assertEqual(bucket_of(20), "10-20")
        self.assertEqual(bucket_of(21), "20-50")
        self.assertEqual(bucket_of(50), "20-50")
        self.assertEqual(bucket_of(51), "50-100")
        self.assertEqual(bucket_of(100), "50-100")
        self.assertEqual(bucket_of(101), "100+")

    def test_unstructured_uses_max_per_person(self):
        counts = {"SSN": 3, "EMAIL": 5, "NAME": 2}
        n = estimate_entities({}, counts, ["SSN", "Email", "Name"], ("SSN", "EMAIL", "NAME"))
        self.assertEqual(n, 5)

    def test_structured_uses_rows(self):
        n = estimate_entities({"is_structured": True, "structured_entity_rows": 60}, {}, [], ())
        self.assertEqual(n, 60)

    def test_floor_when_keyword_only(self):
        # a keyword hit with no per-person identifier still counts as >=1 entity
        self.assertEqual(estimate_entities({}, {"SSN": 0}, ["Health"], ("SSN",)), 1)


class TestRulepackLegacy(unittest.TestCase):
    def test_legacy_patterns_convert(self):
        legacy = {"name": "old",
                  "patterns": {"SSN": r"\d{3}-\d{2}-\d{4}"},
                  "keywords": {"PASSPORT": r"passport"},
                  "per_person_types": ["SSN"]}
        with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False) as fh:
            json.dump(legacy, fh)
            path = fh.name
        try:
            pack = load_rulepack(path)
            keys = {e["key"] for e in pack["entities"]}
            self.assertIn("SSN", keys)
            self.assertIn("PASSPORT", keys)
        finally:
            os.remove(path)


class TestApplyOcr(unittest.TestCase):
    def setUp(self):
        self.cfg = Config(root=".")

    def test_ocr_runs_on_image_only(self):
        def fake(path, cfg):
            return "John Smith SSN 123-45-6789", {"text_extractable": "text"}
        text, meta = apply_ocr("x.pdf", "", {"text_extractable": "image_only"}, self.cfg, fake)
        self.assertIn("123-45-6789", text)
        self.assertEqual(meta["text_extractable"], "text")
        self.assertTrue(meta["ocr"])

    def test_ocr_skips_searchable(self):
        calls = []
        def fake(path, cfg):
            calls.append(1)
            return "x", {}
        text, _ = apply_ocr("x.pdf", "already text", {"text_extractable": "text"}, self.cfg, fake)
        self.assertEqual(text, "already text")
        self.assertEqual(calls, [])

    def test_ocr_none_is_noop(self):
        _, meta = apply_ocr("x.pdf", "", {"text_extractable": "image_only"}, self.cfg, None)
        self.assertEqual(meta["text_extractable"], "image_only")

    def test_ocr_failure_degrades(self):
        def boom(path, cfg):
            raise RuntimeError("nope")
        _, meta = apply_ocr("x.pdf", "", {"text_extractable": "image_only"}, self.cfg, boom)
        self.assertEqual(meta["text_extractable"], "image_only")  # falls back, not dropped
        self.assertIn("ocr_failed", meta["detail"])


class TestApplyLlm(unittest.TestCase):
    def setUp(self):
        self.cfg = Config(root=".")

    def _rec(self, ambiguity, **kw):
        base = dict(rel_path="a", file_name="a", ext=".txt", size_bytes=1,
                    searchable=True, ambiguity=ambiguity)
        base.update(kw)
        return FileRecord(**base)

    def test_llm_only_fires_on_ambiguous(self):
        calls = []
        def fake(text, cfg):
            calls.append(1)
            return {"responsive": True}
        rec = self._rec("clear_responsive")
        apply_llm(rec, "text", self.cfg, fake)
        self.assertEqual(calls, [])
        self.assertFalse(rec.llm_consulted)

    def test_llm_sets_responsive_and_names(self):
        def fake(text, cfg):
            return {"responsive": True, "names": ["a", "b"], "person_count": 2}
        rec = self._rec("ambiguous", estimated_entities=0, entities_found="Email")
        apply_llm(rec, "text", self.cfg, fake)
        self.assertTrue(rec.llm_consulted)
        self.assertEqual(rec.llm_responsive, "yes")
        self.assertEqual(rec.estimated_entities, 2)
        self.assertIn("Name", rec.entities_found)
        self.assertEqual(choose_lane(rec), "standard")

    def test_llm_non_responsive_clears_ambiguous_file(self):
        # the AI trims over-calls by clearing AMBIGUOUS files (weak signals, no strong ID)
        def fake(text, cfg):
            return {"responsive": False}
        rec = self._rec("ambiguous", estimated_entities=1, entities_found="Email")
        apply_llm(rec, "text", self.cfg, fake)
        self.assertEqual(rec.llm_responsive, "no")
        self.assertEqual(choose_lane(rec), "likely_non_responsive")

    def test_llm_can_upgrade_a_nothing_file(self):
        # the AI may also flag a file the rules found nothing in
        def fake(text, cfg):
            return {"responsive": True, "person_count": 2}
        rec = self._rec("ambiguous", estimated_entities=0, entities_found="")
        apply_llm(rec, "text", self.cfg, fake)
        self.assertEqual(choose_lane(rec), "standard")

    def test_llm_none_is_noop(self):
        rec = self._rec("ambiguous", entities_found="Email")
        apply_llm(rec, "text", self.cfg, None)
        self.assertFalse(rec.llm_consulted)

    def test_llm_failure_degrades(self):
        def boom(text, cfg):
            raise RuntimeError("nope")
        rec = self._rec("ambiguous", entities_found="Email")
        apply_llm(rec, "text", self.cfg, boom)
        self.assertFalse(rec.llm_consulted)
        self.assertIn("llm_failed", rec.detail)

    def test_llm_malformed_person_count_falls_back(self):
        # A non-numeric person_count must not crash, nor discard the result; it
        # falls back to the number of names returned.
        def fake(text, cfg):
            return {"responsive": True, "names": ["a", "b"], "person_count": "several"}
        rec = self._rec("ambiguous", estimated_entities=0, entities_found="Email")
        apply_llm(rec, "text", self.cfg, fake)
        self.assertTrue(rec.llm_consulted)
        self.assertEqual(rec.llm_responsive, "yes")
        self.assertEqual(rec.estimated_entities, 2)  # len(names)


class TestProcessFileEnrichment(unittest.TestCase):
    """End-to-end through process_file: an image file is OCR'd into ambiguous text,
    the LLM is consulted, and its call drives the lane -- using injected fakes."""
    def test_ocr_then_llm(self):
        import pii_triage.azure_clients as az
        from pii_triage import runner
        d = tempfile.mkdtemp()
        png = os.path.join(d, "scan.png")
        open(png, "wb").close()  # x_image returns image_only without decoding
        cfg = Config(root=d, use_ocr=True, use_llm=True)
        orig_ocr, orig_llm = az.get_ocr_fn, az.get_llm_fn
        az.get_ocr_fn = lambda c: (lambda path, c2: (
            "Please contact a@b.com about the matter",
            {"text_extractable": "text", "is_structured": False, "page_or_sheet_count": 1}))
        az.get_llm_fn = lambda c: (lambda text, c2: {
            "responsive": True, "names": ["Jane Roe"], "person_count": 1})
        try:
            runner._init_worker(cfg)
            rec = runner.process_file(png)
        finally:
            az.get_ocr_fn, az.get_llm_fn = orig_ocr, orig_llm
            runner._init_worker(Config(root="."))  # reset worker globals
        self.assertTrue(rec["searchable"])            # OCR turned it searchable
        self.assertEqual(rec["ambiguity"], "ambiguous")
        self.assertTrue(rec["llm_consulted"])         # ambiguous -> LLM fired
        self.assertEqual(rec["llm_responsive"], "yes")
        self.assertEqual(rec["suggested_lane"], "standard")


class TestCliWiring(unittest.TestCase):
    def test_get_extractor_imported_in_cli(self):
        # Regression: cli.py used get_extractor without importing it, so any
        # scan with --protocol raised NameError before reading the protocol.
        import pii_triage.cli as cli
        self.assertTrue(callable(getattr(cli, "get_extractor", None)))


class TestDotenv(unittest.TestCase):
    def _write(self, text):
        p = os.path.join(tempfile.mkdtemp(), ".env")
        with open(p, "w", encoding="utf-8") as fh:
            fh.write(text)
        return p

    def test_parses_basic_quotes_comments_export(self):
        keys = ("PIIT_A", "PIIT_B", "PIIT_C")
        for k in keys:
            os.environ.pop(k, None)
        p = self._write("# a comment\n\nPIIT_A=hello\n"
                        'PIIT_B="quoted value"\nexport PIIT_C=exported\n')
        try:
            n = load_dotenv(p)
            self.assertEqual(n, 3)
            self.assertEqual(os.environ["PIIT_A"], "hello")
            self.assertEqual(os.environ["PIIT_B"], "quoted value")
            self.assertEqual(os.environ["PIIT_C"], "exported")
        finally:
            for k in keys:
                os.environ.pop(k, None)

    def test_does_not_override_real_env(self):
        os.environ["PIIT_X"] = "real"
        p = self._write("PIIT_X=fromfile\n")
        try:
            load_dotenv(p)
            self.assertEqual(os.environ["PIIT_X"], "real")        # real env wins
            load_dotenv(p, override=True)
            self.assertEqual(os.environ["PIIT_X"], "fromfile")    # explicit override
        finally:
            os.environ.pop("PIIT_X", None)

    def test_absent_file_is_zero(self):
        self.assertEqual(load_dotenv("/no/such/path/.env"), 0)


class TestBenchmark(unittest.TestCase):
    def _inv(self, rows):
        from pii_triage.routing import FIELDNAMES
        p = os.path.join(tempfile.mkdtemp(), "inv.csv")
        with open(p, "w", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=FIELDNAMES)
            w.writeheader()
            for r in rows:
                base = {k: "" for k in FIELDNAMES}
                base.update(r)
                w.writerow(base)
        return p

    def _score(self, inv, gold, **kw):
        import contextlib
        import io
        with contextlib.redirect_stderr(io.StringIO()):
            return run_benchmark(inv, gold, **kw)

    def test_parse_resp(self):
        for t in ("Responsive", "R", "1", "yes"):
            self.assertTrue(_parse_resp(t))
        for f in ("Non-Responsive", "NR", "0", "no"):
            self.assertFalse(_parse_resp(f))
        self.assertIsNone(_parse_resp(""))
        self.assertIsNone(_parse_resp("Privileged"))

    def test_pick_column(self):
        headers = ["Document Name", "Responsiveness", "Reviewer"]
        self.assertEqual(_pick_column(headers, None, _ID_HINTS), "Document Name")
        self.assertEqual(_pick_column(headers, None, _RESP_HINTS), "Responsiveness")
        self.assertEqual(_pick_column(headers, "Reviewer", _ID_HINTS), "Reviewer")

    def test_pred_respects_ai_lane(self):
        # rules found an email but the AI cleared it -> lane says non-responsive
        self.assertFalse(_pred_from_lane({"suggested_lane": "likely_non_responsive",
                                          "entities_found": "Email"}))
        self.assertTrue(_pred_from_lane({"suggested_lane": "standard"}))
        self.assertIsNone(_pred_from_lane({"suggested_lane": "nonsearchable_sample"}))

    def test_csv_gold_ai_clear_is_agreement(self):
        inv = self._inv([
            {"rel_path": "a.pdf", "file_name": "a.pdf", "suggested_lane": "standard",
             "entities_found": "SSN", "estimated_entities": "3"},
            {"rel_path": "b.pdf", "file_name": "b.pdf",
             "suggested_lane": "likely_non_responsive",
             "entities_found": "Email", "estimated_entities": "1"},  # AI cleared it
        ])
        gold = os.path.join(tempfile.mkdtemp(), "gold.csv")
        with open(gold, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["rel_path", "gold_responsive"])
            w.writerow(["a.pdf", "Responsive"])
            w.writerow(["b.pdf", "Non-Responsive"])
        rep = self._score(inv, gold)
        self.assertEqual((rep["tp"], rep["tn"], rep["fp"], rep["fn"]), (1, 1, 0, 0))

    def test_xlsx_gold_basename_match_and_miss(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        inv = self._inv([
            {"rel_path": "sub/Doc1.pdf", "file_name": "Doc1.pdf", "suggested_lane": "standard"},
            {"rel_path": "sub/Doc2.pdf", "file_name": "Doc2.pdf",
             "suggested_lane": "likely_non_responsive"},
        ])
        gold = os.path.join(tempfile.mkdtemp(), "gold.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Document Name", "Responsiveness"])
        ws.append(["Doc1.pdf", "Responsive"])
        ws.append(["Doc2.pdf", "Responsive"])  # coded responsive; tool cleared -> miss
        wb.save(gold)
        rep = self._score(inv, gold)
        self.assertEqual(rep["matched"], 2)
        self.assertEqual(rep["tp"], 1)
        self.assertEqual(rep["fn"], 1)
        self.assertIn("Doc2.pdf", rep["misses"])

    def test_unmatched_is_reported(self):
        inv = self._inv([{"rel_path": "x.pdf", "file_name": "x.pdf", "suggested_lane": "standard"}])
        gold = os.path.join(tempfile.mkdtemp(), "gold.csv")
        with open(gold, "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["filename", "responsive"])
            w.writerow(["not_in_scan.pdf", "Responsive"])
        rep = self._score(inv, gold)
        self.assertEqual(rep["unmatched"], 1)
        self.assertIn("not_in_scan.pdf", rep["unmatched_files"])

    def test_two_sheet_targets_manual_review_column(self):
        # Mirrors the real file: a model-prediction column ("Responsive?") AND the
        # human truth ("Manual Review Responsive"), across two sheets.
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")
        inv = self._inv([
            {"rel_path": "q1.pdf", "file_name": "q1.pdf", "suggested_lane": "standard"},
            {"rel_path": "q2.pdf", "file_name": "q2.pdf", "suggested_lane": "likely_non_responsive"},
        ])
        gold = os.path.join(tempfile.mkdtemp(), "results.xlsx")
        wb = openpyxl.Workbook()
        first = wb.active
        first.title = "GPT 5 Results"
        first.append(["File Name", "Responsive?", "Manual Review Responsive"])
        first.append(["q1.pdf", "FALSE", "FALSE"])   # decoy sheet: opposite truth
        second = wb.create_sheet("GPT 5.4 Nano Results")
        second.append(["File Name", "Responsive?", "Manual Review Responsive"])
        second.append(["q1.pdf", "FALSE", "TRUE"])   # model said FALSE, human says TRUE
        second.append(["q2.pdf", "TRUE", "FALSE"])   # model said TRUE, human says FALSE
        wb.save(gold)
        rep = self._score(inv, gold, responsive_col="Manual Review Responsive",
                          sheet="GPT 5.4 Nano Results")
        self.assertEqual(rep["gold_sheet"], "GPT 5.4 Nano Results")
        self.assertEqual(rep["gold_columns"]["responsive"], "Manual Review Responsive")
        # q1: human TRUE, tool responsive -> TP ; q2: human FALSE, tool cleared -> TN
        self.assertEqual((rep["tp"], rep["tn"], rep["fp"], rep["fn"]), (1, 1, 0, 0))


class TestQuietWarnings(unittest.TestCase):
    def test_pdfminer_logger_quieted(self):
        import logging
        from pii_triage.extractors import quiet_noisy_libraries
        logging.getLogger("pdfminer").setLevel(logging.WARNING)
        quiet_noisy_libraries()
        self.assertEqual(logging.getLogger("pdfminer").level, logging.ERROR)


class TestSsnFormats(unittest.TestCase):
    def test_dashed_and_spaced(self):
        self.assertEqual(detect_ssn("ssn 123-45-6789"), 1)
        self.assertEqual(detect_ssn("SSN: 123 45 6789"), 1)

    def test_labeled_bare_nine_digits(self):
        # the real-world miss: SSN written as 9 digits next to its label
        self.assertEqual(detect_ssn("Social Security Number: 048021413"), 1)
        self.assertEqual(detect_ssn("whose social security number is 030604131"), 1)

    def test_bare_without_label_ignored(self):
        self.assertEqual(detect_ssn("Order number 048021413 shipped"), 0)
        self.assertEqual(detect_ssn("Account 123456789 balance"), 0)

    def test_ein_and_charity_not_ssn(self):
        self.assertEqual(detect_ssn("EIN 36-2193610"), 0)
        self.assertEqual(detect_ssn("CRA #72740 9880 RR0001"), 0)

    def test_invalid_area_filtered(self):
        self.assertEqual(detect_ssn("ssn 000-12-3456"), 0)
        self.assertEqual(detect_ssn("ssn 666-12-3456"), 0)

    def test_new_hire_form_is_clear_responsive(self):
        text = "Social Security Number: 048021413\nEMPLOYEE NAME: Jeffrey Meyers"
        counts, labels, _ = detect(text, RULES)
        self.assertEqual(counts["SSN"], 1)
        self.assertEqual(classify_ambiguity(counts, labels, False), "clear_responsive")


class TestValueNotMention(unittest.TestCase):
    """Identifier types require an actual value next to the label, not a bare
    mention -- this is the false-responsive (over-call) fix."""

    def test_mentions_do_not_flag(self):
        for t in ["please bring your passport",
                  "enter your account number below",
                  "we may collect your driver's license",
                  "your date of birth field",
                  "Passport Office downtown",
                  "tax identification guidelines"]:
            _, labels, _ = detect(t, RULES)
            self.assertEqual(labels, [], f"over-flagged a mention: {t!r} -> {labels}")

    def test_real_values_flag(self):
        cases = {
            "Passport No: X1234567": "Passport",
            "Account number: 123456789": "Bank/Account Number",
            "DL No: D1234567": "Driver License",
            "Date of Birth: 02/09/2001": "Date of Birth",
            "Tax ID 12-3456789": "Tax ID",
        }
        for t, lab in cases.items():
            _, labels, _ = detect(t, RULES)
            self.assertIn(lab, labels, f"missed a real value: {t!r} -> {labels}")

    def test_blank_form_fields_not_flagged(self):
        _, labels, _ = detect("Date of Birth: ____   Passport No: ____", RULES)
        self.assertEqual(labels, [])


class TestNoiseSilencing(unittest.TestCase):
    def test_pypdf_logger_quieted(self):
        import logging
        from pii_triage.extractors import quiet_noisy_libraries
        quiet_noisy_libraries()
        # the 'Impossible to decode XFormObject' line comes from pypdf._page
        self.assertGreaterEqual(
            logging.getLogger("pypdf._page").getEffectiveLevel(), logging.ERROR)


class TestSsnSeparation(unittest.TestCase):
    def test_label_and_value_separated_still_detected(self):
        # form-field / OCR style: the SSN label and the number land far apart
        text = ("Social Security Number\nEmployee Name\nStreet Address\n"
                + "filler " * 30 + "\n048021413")
        self.assertEqual(detect_ssn(text), 1)

    def test_no_label_still_ignored(self):
        text = "filler " * 30 + " 048021413 order confirmation"
        self.assertEqual(detect_ssn(text), 0)


try:
    from reportlab.pdfgen import canvas as _rl_canvas
    _HAVE_RL = True
except Exception:
    _HAVE_RL = False

try:
    from PIL import Image as _PILImage
    _HAVE_PIL = True
except Exception:
    _HAVE_PIL = False


class TestPdfFormExtraction(unittest.TestCase):
    @unittest.skipUnless(_HAVE_RL, "reportlab not installed")
    def test_fillable_form_values_detected(self):
        """A typed value in a PDF form FIELD is invisible to page-text extraction -- it lives
        in the AcroForm dictionary. _pdf_field_values pulls it in as "field name: value", which
        also supplies the label that value-near-label rules need.

        3.0.0: this test previously built the PDF and asserted NOTHING, so it passed
        unconditionally while verifying nothing. A fillable form carrying an SSN is exactly the
        document that must never be cleared, so it now checks the whole chain: extraction,
        then detection.
        """
        import tempfile, os
        from pii_triage.config import Config
        from pii_triage.extractors import x_pdf
        from pii_triage.detection import detect
        p = os.path.join(tempfile.gettempdir(), "rl_fill_test.pdf")
        c = _rl_canvas.Canvas(p)
        c.drawString(72, 700, "Social Security Number:")
        c.acroForm.textfield(name="Social Security Number", value="048021413",
                             x=200, y=695, width=180, height=14)
        c.acroForm.textfield(name="E-Mail Address", value="a.person@example.org",
                             x=200, y=675, width=180, height=14)
        c.save()

        text, meta = x_pdf(p, Config(root=tempfile.gettempdir()), RULES)
        # the field VALUES must reach the extracted text, not just the page label
        self.assertIn("048021413", text, "SSN in a form field was not extracted")
        self.assertIn("a.person@example.org", text, "email in a form field was not extracted")
        self.assertEqual(meta["text_extractable"], "text")
        # ...and detection must see them
        counts, labels, _cats = detect(text, RULES)
        self.assertEqual(counts.get("SSN"), 1, f"SSN not detected; labels={labels}")
        self.assertGreaterEqual(counts.get("EMAIL", 0), 1, f"email not detected; labels={labels}")
        self.assertIn("SSN", labels)

    @unittest.skipUnless(_HAVE_RL and _HAVE_PIL, "reportlab/PIL not installed")
    def test_image_dominated_pdf_routes_to_ocr(self):
        # big bytes + little text, with an image too short to trip the full-page
        # check -- must still go to OCR via the bytes-per-character signal.
        import tempfile, os
        from pii_triage.config import Config
        from pii_triage.extractors import x_pdf
        w, h = 1400, 400                       # h < 550 so the full-page check is off
        img = os.path.join(tempfile.gettempdir(), "rl_wide_noise.jpg")
        _PILImage.frombytes("RGB", (w, h), os.urandom(w * h * 3)).save(img, quality=95)
        p = os.path.join(tempfile.gettempdir(), "rl_bigscan.pdf")
        c = _rl_canvas.Canvas(p)
        c.drawImage(img, 30, 400, width=540, height=160)
        c.drawString(60, 60, "form")           # tiny text scrap
        c.save()
        _, meta = x_pdf(p, Config(root=tempfile.gettempdir()), RULES)
        self.assertEqual(meta["text_extractable"], "image_only")

    @unittest.skipUnless(_HAVE_RL and _HAVE_PIL, "reportlab/PIL not installed")
    def test_letterhead_logo_does_not_trigger_ocr(self):
        import tempfile, os, pypdf
        from pii_triage.extractors import _pdf_image_based
        logo = os.path.join(tempfile.gettempdir(), "rl_logo.png")
        _PILImage.new("RGB", (150, 60), "blue").save(logo)
        p = os.path.join(tempfile.gettempdir(), "rl_letterhead.pdf")
        c = _rl_canvas.Canvas(p)
        c.drawImage(logo, 60, 740, width=120, height=48)        # small logo
        c.drawString(60, 700, "An ordinary letter with real text content. " * 8)
        c.showPage()
        c.save()
        rdr = pypdf.PdfReader(p)
        self.assertFalse(_pdf_image_based(rdr, len(rdr.pages)))


class TestStderrHush(unittest.TestCase):
    def test_hush_silences_stderr(self):
        import io, sys, contextlib
        from pii_triage.extractors import _hush_stderr
        outer = io.StringIO()
        with contextlib.redirect_stderr(outer):
            with _hush_stderr():
                sys.stderr.write("Impossible to decode XFormObject /X4: '/XObject'")
        self.assertEqual(outer.getvalue(), "")


class TestLockedOutputFile(unittest.TestCase):
    def test_locked_csv_gives_friendly_exit(self):
        import io, contextlib, tempfile, os
        from unittest import mock
        from pii_triage.config import Config
        from pii_triage import runner
        out = os.path.join(tempfile.gettempdir(), "lock_unit_test.csv")
        open(out, "w").write("x")
        err = io.StringIO()
        with mock.patch.object(runner.os, "remove",
                               side_effect=PermissionError(32, "in use")):
            with self.assertRaises(SystemExit) as cm:
                with contextlib.redirect_stderr(err):
                    runner.run(Config(root=tempfile.gettempdir()),
                               [], out, 1, 999.0, 1, True)
        self.assertEqual(cm.exception.code, 2)
        self.assertIn("open in another program", err.getvalue())



class TestCardValidation(unittest.TestCase):
    def test_real_cards_pass(self):
        for num in ("4111111111111111",      # Visa
                    "5555555555554444",      # Mastercard
                    "378282246310005",       # Amex
                    "6011111111111117"):     # Discover
            self.assertTrue(card_valid(num), num)

    def test_luhn_valid_order_number_is_not_a_card(self):
        # passes the checksum but starts with a non-issuer prefix -> not a card
        order = "8123456000000009"
        self.assertTrue(luhn_valid(order))
        self.assertFalse(card_valid(order))

    def test_card_not_detected_for_bare_order_numbers(self):
        text = "Order 8123456000000009 shipped; ref 7000000000000007 invoice 9000000000000001"
        counts, labels, _ = detect(text, RULES)
        self.assertEqual(counts.get("CARD", 0), 0)
        self.assertNotIn("Payment Card", labels)

    def test_card_still_detected_for_real_card(self):
        counts, labels, _ = detect("Visa 4111 1111 1111 1111 on file", RULES)
        self.assertGreaterEqual(counts.get("CARD", 0), 1)
        self.assertIn("Payment Card", labels)


class TestLlmInputSampler(unittest.TestCase):
    def test_short_text_unchanged(self):
        t = "name: a\nssn label\n" * 20
        self.assertEqual(_sample_for_llm(t, 5000), t)

    def test_long_text_is_bounded_and_spans_file(self):
        big = ("H" * 200) + ("M" * 200_000) + ("T" * 200)
        s = _sample_for_llm(big, 24_000)
        self.assertLessEqual(len(s), 24_000 + 200)   # +markers
        self.assertTrue(s.startswith("H"))           # keeps the head (headers/first rows)
        self.assertTrue(s.endswith("T"))             # keeps the tail
        self.assertIn("middle of file", s)           # samples the middle

    def test_zero_limit_returns_full_text(self):
        self.assertEqual(_sample_for_llm("abc", 0), "abc")


if __name__ == "__main__":
    unittest.main()
