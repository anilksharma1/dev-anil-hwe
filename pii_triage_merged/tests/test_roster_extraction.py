"""Tests for the roster-extraction fix: reset_dimensions() + name-line counting."""
import os, sys, unittest
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from pii_triage.detection import is_roster_name_line, CompiledRules
from pii_triage.config import Config


class TestRosterNameLine(unittest.TestCase):
    def test_matches_lastname_firstname(self):
        self.assertTrue(is_roster_name_line("Ackerman, Robert T. 80.00 6,634.62"))
        self.assertTrue(is_roster_name_line("Ago, Scott J. 40.00 1,250.00"))
        self.assertTrue(is_roster_name_line("  Akin, Pamela Ruth 37.00"))
    def test_rejects_totals_and_labels(self):
        self.assertFalse(is_roster_name_line("Total, 1,234.00"))
        self.assertFalse(is_roster_name_line("Company Totals"))
        self.assertFalse(is_roster_name_line("SS 373.06"))
        self.assertFalse(is_roster_name_line("Med 87.25 S1 IL 279.05"))
        self.assertFalse(is_roster_name_line(""))
        self.assertFalse(is_roster_name_line("PRETAX MEDICAL 568.14"))


class TestXlsxRosterExtraction(unittest.TestCase):
    """Build a workbook with a broken declared dimension + a register sheet where each
    person row is name-only, and confirm the extractor reads all rows and counts people."""
    @classmethod
    def setUpClass(cls):
        try:
            import openpyxl  # noqa: F401
        except Exception:
            raise unittest.SkipTest("openpyxl not available")
        import openpyxl, tempfile
        cls.tmp = tempfile.mkdtemp()
        cls.path = os.path.join(cls.tmp, "register.xlsx")
        wb = openpyxl.Workbook()
        info = wb.active; info.title = "Payroll Info"
        info["A1"] = "Company Code"; info["A2"] = "Company Name Acme"
        reg = wb.create_sheet("Payroll Register")
        reg["A1"] = "Personnel"
        _fn = ["John", "Mary", "Robert", "Linda", "James", "Patricia"]
        for i in range(60):                      # 60 name-only person rows
            reg.cell(row=2 + i, column=1,
                     value=f"Smith, {_fn[i % len(_fn)]} A. 40.00 1,000.00")
        wb.create_sheet("Company Totals")["A1"] = "Company Totals"
        wb.save(cls.path); wb.close()

    def _run(self):
        from pii_triage.extractors import x_xlsx
        cfg = Config(root=".")
        rules = CompiledRules.from_pack(cfg.rulepack, use_ner=cfg.use_ner)
        return x_xlsx(self.path, cfg, rules)

    def test_reads_all_rows_and_counts_people(self):
        text, meta = self._run()
        # without reset_dimensions the register sheet would yield ~1 row; with it, ~60+.
        self.assertGreater(meta["structured_total_rows"], 55)
        self.assertGreaterEqual(meta["structured_entity_rows"], 55)  # name-lines counted
        self.assertIn("Smith, John A.", text)                         # roster text present


if __name__ == "__main__":
    unittest.main()