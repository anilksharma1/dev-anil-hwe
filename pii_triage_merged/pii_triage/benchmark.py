"""Accuracy harness: compare a produced inventory against a labeled gold set.

Reads the gold set from .xlsx or .csv. Auto-detects the file-identifier column
and the responsive/NR column (override with explicit names if needed), matches
each gold row to the inventory by relative path or by file name, and scores the
tool's ACTUAL responsiveness decision (its routing lane -- which already reflects
the AI's call on ambiguous files), not rules alone.

The metric that matters most is recall on "responsive": a miss = a missed
notification. The report lists every disagreement by file name so they can be
inspected.
"""
from __future__ import annotations

import csv
import os
import sys

from .routing import bucket_of

# Header hints for auto-detection (checked exact-first, then as a substring).
_ID_HINTS = ("rel_path", "relpath", "file_name", "filename", "file name", "file",
             "document name", "doc name", "docname", "document", "control number",
             "control", "bates", "begdoc", "beg doc", "name", "path")
_RESP_HINTS = ("gold_responsive", "responsive", "responsiveness", "responsive_nr",
               "responsive/nr", "responsive nr", "resp", "determination",
               "coding", "decision", "nr")
_BDE_HINTS = ("gold_bde", "bde", "big data", "51+", "big data extraction")

_TRUE = {"responsive", "r", "yes", "y", "1", "true", "t", "resp", "responsive (r)"}
_FALSE = {"non-responsive", "nonresponsive", "non responsive", "not responsive",
          "nr", "no", "n", "0", "false", "f", "non-responsive (nr)"}

_RESP_LANES = ("standard", "bde", "structured_bde")


def _read_rows(path: str, sheet: str = None):
    """Return (headers, list-of-dict-rows, sheet_name) from a .csv/.tsv or
    .xlsx/.xlsm file. The first row is the header row. For workbooks, `sheet`
    selects a tab by name (default: the active/last-selected tab)."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet:
            if sheet not in wb.sheetnames:
                raise SystemExit(f"error: sheet '{sheet}' not found. Tabs are: {wb.sheetnames}")
            ws = wb[sheet]
        else:
            ws = wb.active
        used = ws.title
        headers, rows = None, []
        for r in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [("" if c is None else str(c)).strip() for c in r]
                continue
            rows.append({h: ("" if v is None else str(v))
                         for h, v in zip(headers, r)})
        wb.close()
        return headers or [], rows, used
    delim = "\t" if ext == ".tsv" else ","
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:  # -sig: strip Excel BOM
        reader = csv.DictReader(fh, delimiter=delim)
        return (reader.fieldnames or []), list(reader), None


def _pick_column(headers, override, hints):
    if override:
        for h in headers:
            if h.strip().lower() == override.strip().lower():
                return h
        raise SystemExit(f"error: column '{override}' not found. Columns are: {headers}")
    low = [(h, h.strip().lower()) for h in headers]
    for h, l in low:                       # exact hint match wins
        if l in hints:
            return h
    for hint in hints:                     # then first header that contains a hint
        for h, l in low:
            if hint in l:
                return h
    return None


def _parse_resp(v):
    s = str(v).strip().lower()
    if s in _TRUE:
        return True
    if s in _FALSE:
        return False
    # An entity-count column ("total entities" = 315, 53, 0, ...): >0 entities -> responsive.
    # Common when the review sheet reports per-file counts rather than a Responsive/NR label.
    try:
        return float(s) > 0
    except ValueError:
        return None  # blank / "privileged" / unknown -> not scored


def _pred_from_lane(rec):
    """The tool's actual responsiveness call (lane already reflects the AI override).
    Returns True/False, or None when the tool made no call (file needs prep/review)."""
    lane = (rec.get("suggested_lane") or "").strip()
    if lane in _RESP_LANES:
        return True
    if lane == "likely_non_responsive":
        return False
    return None


def _reason(rec) -> str:
    """A short, non-PII explanation of why the tool called a file responsive: the
    lane it took, whether the AI was consulted (and its verdict), what labels were
    found, and whether it was treated as a spreadsheet. Labels/status only -- no values."""
    def _t(v):
        return str(v).strip().lower() in ("true", "1", "yes", "y")
    lane = rec.get("suggested_lane") or "-"
    found = rec.get("entities_found") or "(nothing)"
    ai = (f"AI=yes->{rec.get('llm_responsive') or '?'}" if _t(rec.get("llm_consulted"))
          else "AI=not used")
    tag = " | structured" if _t(rec.get("is_structured")) else ""
    return f"lane={lane} | {ai} | found={found}{tag}"


def _prf(tp, fp, fn):
    p = tp / (tp + fp) if (tp + fp) else 0.0
    r = tp / (tp + fn) if (tp + fn) else 0.0
    f1 = 2 * p * r / (p + r) if (p + r) else 0.0
    return p, r, f1


def run_benchmark(inventory_csv: str, gold_path: str, id_col: str = None,
                  responsive_col: str = None, bde_col: str = None,
                  sheet: str = None, absent_means: str = "unreviewed",
                  bde_threshold: int = None) -> dict:
    """absent_means controls inventory files the gold never mentions:
       'unreviewed' (default) -> not scored; 'zero' -> scored as non-responsive
       (assume zero entities), so the tool over-flagging one is a false positive.
       bde_threshold, if set, re-classifies BDE from the file's recorded entity count
       (estimated_entities >= threshold) instead of the run's own is_bde flag -- so a
       run scanned at one threshold can be scored at another (7+ vs 51+)."""
    _, inv_rows, _ = _read_rows(inventory_csv)
    by_rel, by_base = {}, {}
    for r in inv_rows:
        rel = (r.get("rel_path") or "").strip()
        if rel:
            by_rel[rel] = r
        base = (r.get("file_name") or os.path.basename(rel)).strip().lower()
        if base:
            by_base.setdefault(base, r)  # first occurrence wins

    g_headers, g_rows, g_sheet = _read_rows(gold_path, sheet=sheet)
    idc = _pick_column(g_headers, id_col, _ID_HINTS)
    rc = _pick_column(g_headers, responsive_col, _RESP_HINTS)
    bc = _pick_column(g_headers, bde_col, _BDE_HINTS)
    if idc is None or rc is None:
        raise SystemExit(
            "error: couldn't identify the gold columns automatically.\n"
            f"  columns found: {g_headers}\n"
            "  re-run with --id-col \"<file column>\" --responsive-col \"<responsive column>\"")

    tp = fp = fn = tn = 0
    bde_correct = bde_total = 0
    matched = undetermined = unmatched = unscored = 0
    misses, overcalls, unmatched_files = [], [], []
    miss_reasons, overcall_reasons = {}, {}
    matched_ids = set()   # inventory records consumed by a gold row (for absent_means="zero")

    for g in g_rows:
        gold_resp = _parse_resp(g.get(rc, ""))
        if gold_resp is None:
            unscored += 1
            continue
        key = str(g.get(idc, "")).strip()
        rec = (by_rel.get(key) or by_base.get(key.lower())
               or by_base.get(os.path.basename(key).lower()))
        if rec is None:
            unmatched += 1
            unmatched_files.append(key)
            continue
        matched += 1
        matched_ids.add(id(rec))
        pred = _pred_from_lane(rec)
        if pred is None:
            undetermined += 1
            continue
        if gold_resp and pred:
            tp += 1
        elif gold_resp and not pred:
            fn += 1
            fname = rec.get("file_name") or key
            misses.append(fname)
            miss_reasons[fname] = _reason(rec)
        elif not gold_resp and pred:
            fp += 1
            fname = rec.get("file_name") or key
            overcalls.append(fname)
            overcall_reasons[fname] = _reason(rec)
        else:
            tn += 1
        if bc and str(g.get(bc, "")).strip() != "":
            bde_total += 1
            gv = str(g.get(bc)).strip()
            if bde_threshold is not None:
                # count-based BDE: compare BOTH the tool's count and the gold's count to the same
                # threshold (so a "total entities" column scores directly as 51+/7+).
                try:
                    pred_bde = int(float(rec.get("estimated_entities") or 0)) >= bde_threshold
                except (TypeError, ValueError):
                    pred_bde = False
                try:
                    gold_bde = float(gv) >= bde_threshold
                except ValueError:
                    gold_bde = gv.lower() in ("1", "true", "yes", "y", "bde")
            else:
                pred_bde = str(rec.get("is_bde")).strip().lower() in ("true", "1")
                gold_bde = gv.lower() in ("1", "true", "yes", "y", "bde")
            if pred_bde == gold_bde:
                bde_correct += 1

    # "assume zero entities on files not in the manual review": score every inventory file the gold
    # never mentioned as non-responsive (opt-in). The tool over-flagging an unreviewed file then
    # counts as a false positive; clearing it counts as a true negative. Files the tool couldn't call
    # (needs conversion/parser/OCR) are still skipped.
    assumed_nr = 0
    if absent_means == "zero":
        for rec in inv_rows:
            if id(rec) in matched_ids:
                continue
            pred = _pred_from_lane(rec)
            if pred is None:
                continue
            assumed_nr += 1
            if pred:
                fp += 1
                fname = rec.get("file_name") or rec.get("rel_path") or "?"
                overcalls.append(fname)
                overcall_reasons[fname] = _reason(rec)
            else:
                tn += 1

    precision, recall, f1 = _prf(tp, fp, fn)
    nr_precision, nr_recall, nr_f1 = _prf(tn, fn, fp)
    report = {
        "gold_columns": {"id": idc, "responsive": rc, "bde": bc},
        "gold_sheet": g_sheet,
        "gold_rows": len(g_rows), "scored": tp + fp + fn + tn,
        "matched": matched, "undetermined": undetermined,
        "unmatched": unmatched, "unscored_blank_label": unscored,
        "absent_means": absent_means, "assumed_nr": assumed_nr,
        "responsive_precision": round(precision, 4),
        "responsive_recall": round(recall, 4),
        "responsive_f1": round(f1, 4),
        "nr_precision": round(nr_precision, 4),
        "nr_recall": round(nr_recall, 4),
        "nr_f1": round(nr_f1, 4),
        "tp": tp, "fp": fp, "fn": fn, "tn": tn,
        "bde_accuracy": round(bde_correct / bde_total, 4) if bde_total else None,
        "bde_threshold": bde_threshold,
        "misses": misses, "overcalls": overcalls, "unmatched_files": unmatched_files,
        "miss_reasons": miss_reasons, "overcall_reasons": overcall_reasons,
    }
    _print_report(report)
    return report


def _print_report(r: dict) -> None:
    w = sys.stderr.write
    c = r["gold_columns"]
    agree = r["tp"] + r["tn"]
    w("\n=== pii_triage vs your results ===\n")
    if r.get("gold_sheet"):
        w(f"reading sheet: '{r['gold_sheet']}'\n")
    w(f"using gold columns -> file: '{c['id']}'  responsive: '{c['responsive']}'"
      + (f"  bde: '{c['bde']}'" if c["bde"] else "") + "\n")
    w(f"gold rows: {r['gold_rows']}  scored: {r['scored']}  "
      f"agree: {agree}  disagree: {r['fp'] + r['fn']}\n")
    if r.get("absent_means") == "zero":
        w(f"absent-means=zero: {r.get('assumed_nr', 0)} inventory file(s) not in the review "
          f"scored as non-responsive (assumed zero entities)\n")
    if r["scored"]:
        w(f"agreement            : {agree}/{r['scored']} = {agree / r['scored']:.1%}\n")
    w(f"responsive recall    : {r['responsive_recall']:.3f}   "
      f"(misses = files you coded responsive that the tool cleared)\n")
    w(f"responsive precision : {r['responsive_precision']:.3f}\n")
    w(f"responsive F1        : {r['responsive_f1']:.3f}\n")
    w(f"NR recall            : {r['nr_recall']:.3f}   "
      f"(over-calls = files you coded NR that the tool flagged)\n")
    w(f"NR precision         : {r['nr_precision']:.3f}\n")
    w(f"NR F1                : {r['nr_f1']:.3f}\n")
    w(f"confusion            : TP={r['tp']} FP={r['fp']} FN={r['fn']} TN={r['tn']}\n")
    if r["bde_accuracy"] is not None:
        thr = r.get("bde_threshold")
        w(f"BDE flag accuracy    : {r['bde_accuracy']:.3f}"
          + (f"   (re-scored at {thr}+ entities)\n" if thr is not None else "\n"))
    if r["undetermined"]:
        w(f"note: {r['undetermined']} coded file(s) the tool couldn't make a call on "
          f"(needs conversion/parser/OCR) -- excluded from the scores above.\n")
    if r["unmatched"]:
        w(f"note: {r['unmatched']} coded file(s) weren't found in the scan by name "
          f"(check the file-id column matches the scanned filenames):\n")
        for f in r["unmatched_files"][:15]:
            w(f"    - {f}\n")
    if r["misses"]:
        w(f"\nMISSES ({len(r['misses'])}) -- you coded responsive, tool cleared (inspect first):\n")
        for f in r["misses"][:25]:
            w(f"    - {f}\n      {r.get('miss_reasons', {}).get(f, '')}\n")
    if r["overcalls"]:
        w(f"\nOVER-CALLS ({len(r['overcalls'])}) -- tool flagged, you coded non-responsive:\n")
        for f in r["overcalls"][:30]:
            w(f"    - {f}\n      {r.get('overcall_reasons', {}).get(f, '')}\n")
