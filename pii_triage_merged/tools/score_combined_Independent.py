#!/usr/bin/env python3
r"""
score_combined_Independent.py -- standalone fork of score_combined.py.

Same scoring logic, but with zero dependency on the rest of this repo (no
python-dotenv, no pii_triage package, no scaling_lib requirement) so it can
be copied out and run locally against the outputs of an already-completed
run. scaling-lib timing data is still used opportunistically if it happens
to be importable, but its absence never blocks scoring.

score_combined.py -- one scorecard for a pii_triage 3.0.0 combined-run inventory.

Replaces running score_bde_by_count.py and metrics_vs_manual.py separately, and adds the
things the single-pass build made measurable for the first time: OCR/Document Intelligence
spend, per-stage LLM spend, and stage-2 accuracy alongside stage 1.

WHAT IT SCORES
  1. RUN COST        DI calls/pages and per-stage tokens, priced, plus the two-pass
                     counterfactual (what the old Anna-then-Daniel workflow would have cost).
  2. NR/R ACCURACY   stage 1 (Anna's), stage 2 (Daniel's), the SEQUENTIAL PIPELINE verdict,
                     and the UNION -- each on its correct population.
  3. BDE ACCURACY    four definitions (as in score_bde_by_count.py) plus stage 2's own flag,
                     plus count accuracy against the human count.
  4. BREAKDOWNS      by file type, by searchable/structured, and the OCR-yield question.
  5. TIMING          per-stage wall times and throughput from scaling-lib (when available).

GROUND TRUTH -- the entities export does double duty
  The CNG entities export gives a per-file "Total Entities" count, so ONE file supplies both
  truths:
      responsive  <=>  Total Entities > 0     ("zero entities means no PII")
      BDE         <=>  Total Entities > --bde-threshold   (default 6, i.e. 7+)

  What does it mean when a file is in the inventory but NOT in the entities export?
    --absent-means zero        the export lists only files that HAVE entities, so absent = 0
                               entities = genuinely non-responsive. Scores the whole corpus.
    --absent-means unreviewed  the export lists every reviewed file including zero-count rows,
                               so absent = never reviewed. Scores only the intersection.
    --absent-means auto        (default) infers from whether the export contains any 0 rows.

OUTPUT
  scorecard_YYYYMMDD.xlsx  in --out-dir.  Sheets:
    Run Info | Cost | Timing | NR/R Accuracy | BDE Accuracy | Metrics by Type |
    Stage 2 Detail | OCR Yield | File Detail | Per-file Timing | Misses

  Requires openpyxl (pip install openpyxl); falls back to CSV-only if missing.

USAGE
  python tools/score_combined.py --inventory inventory.csv \
      --entities "CNG_Entities Export.csv" \
      --price-per-1k-in 0.10 --price-per-1k-out 0.40 --price-per-1k-pages 10.0

Stdlib only for reading; openpyxl for xlsx output. Python 3.10+.
"""
from __future__ import annotations

import argparse
import csv
import datetime
import json
import os
import re
import statistics
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter, defaultdict

# ------------------------------------------------------------------ defaults --
DEFAULT_INVENTORY   = "inventory.csv"
DEFAULT_ID_COL      = "Control ID"
DEFAULT_COUNT_COL   = "Total Entities"
DEFAULT_STATUS_COL  = "Data Entry Status"
DEFAULT_BDE_THRESHOLD = 6   # BDE = count > 6  (7+)

RESP_LANES = {"standard", "bde", "structured_bde"}
NR_LANE    = "likely_non_responsive"
BDE_LANES  = {"bde", "structured_bde"}

EXT_TO_TYPE = {
    ".pdf": "PDF", ".xls": "Excel", ".xlsx": "Excel", ".xlsm": "Excel",
    ".xlsb": "Excel", ".msg": "Messages", ".eml": "Messages", ".doc": "MS Word",
    ".docx": "MS Word", ".txt": "Text", ".csv": "Text", ".ppt": "MS PowerPoint",
    ".pptx": "MS PowerPoint", ".html": "HTML", ".htm": "HTML", ".png": "Images",
    ".jpg": "Images", ".jpeg": "Images", ".tif": "Images", ".tiff": "Images",
    ".gif": "Images", ".bmp": "Images",
}

_FTYPE_ORDER = ["Excel", "HTML", "Images", "MS PowerPoint", "MS Word",
                "Messages", "Other", "PDF", "Text"]

_LEGACY_EXTS = {".doc", ".xls", ".ppt"}   # Windows-worker file types


def type_of(name: str) -> str:
    return EXT_TO_TYPE.get(os.path.splitext(name or "")[1].lower(), "Other")


# Do NOT use os.path.splitext: "Q00897.01-0000000003" has no extension but
# splitext cuts it to "Q00897", collapsing thousands of IDs.
# The lookaheads require at least one alpha in the matched suffix so that purely
# numeric segments like ".0001" or ".01" (part of the Control ID, not an extension)
# are never stripped — real file extensions always contain at least one letter.
_EXT_RE = re.compile(r"\.(?=[A-Za-z0-9]{1,5}$)(?=[A-Za-z0-9]*[A-Za-z])[A-Za-z0-9]{1,5}$")


def norm_id(name) -> str:
    return _EXT_RE.sub("", os.path.basename(str(name or "").strip())).strip().lower()


def _b(v) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes")


def _i(v) -> int:
    try:
        return int(float(str(v).strip() or 0))
    except (TypeError, ValueError):
        return 0


def pct(n, d):
    return (n / d) if d else 0.0


# ------------------------------------------------- stdlib .xlsx column reader --
def _local(tag):
    return tag.split("}")[-1]


def _colletter(ref):
    out = []
    for ch in ref:
        if ch.isalpha():
            out.append(ch)
        else:
            break
    return "".join(out)


def _cellval(c, shared):
    for ch in c:
        if _local(ch.tag) == "is":
            return "".join(x.text or "" for x in ch.iter() if _local(x.tag) == "t")
    v = None
    for ch in c:
        if _local(ch.tag) == "v":
            v = ch.text
    if v is None:
        return ""
    if c.get("t") == "s":
        try:
            return shared[int(v)]
        except (ValueError, IndexError):
            return ""
    return v


def read_xlsx_rows(path, sheet_name=""):
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        shared = []
        if "xl/sharedStrings.xml" in names:
            root = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in root:
                shared.append("".join(t.text or "" for t in si.iter()
                                       if _local(t.tag) == "t"))
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        sheets = []
        for sh in wb.iter():
            if _local(sh.tag) == "sheet":
                rid = next((v for k, v in sh.attrib.items() if _local(k) == "id"), None)
                sheets.append((sh.get("name"), rid))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {r.get("Id"): r.get("Target") for r in rels}
        if sheet_name:
            match = next((n for n, _ in sheets
                          if (n or "").strip().lower() == sheet_name.strip().lower()), None)
            if match is None:
                raise KeyError(f"sheet {sheet_name!r} not found. "
                               f"Sheets: {[n for n, _ in sheets]}")
            name = match
        else:
            name = sheets[0][0]
        target = rid_to_target[dict(sheets)[name]]
        target = target[1:] if target.startswith("/") else (
            target if target.startswith("xl/") else "xl/" + target)
        sheet = ET.fromstring(z.read(target))
    for row in sheet.iter():
        if _local(row.tag) != "row":
            continue
        cells = {}
        for c in row:
            if _local(c.tag) == "c":
                cells[_colletter(c.get("r") or "")] = _cellval(c, shared)
        yield cells


def load_two_columns(path, sheet, col_a, col_b):
    pairs = []
    if path.lower().endswith((".csv", ".tsv", ".txt")):
        delim = "\t" if path.lower().endswith(".tsv") else ","
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
            rd = csv.DictReader(fh, delimiter=delim)
            cmap = {(c or "").strip().lower(): c for c in (rd.fieldnames or [])}
            ka, kb = cmap.get(col_a.lower()), cmap.get(col_b.lower())
            if not ka or not kb:
                sys.exit(f"ERROR: need columns {col_a!r} and {col_b!r}.\n"
                         f"       Found: {rd.fieldnames}")
            for r in rd:
                pairs.append((r.get(ka) or "", r.get(kb) or ""))
    else:
        rows = list(read_xlsx_rows(path, sheet))
        ca = cb = None
        hidx = 0
        for i, cells in enumerate(rows[:25]):
            for col, val in cells.items():
                v = (val or "").strip().lower()
                if v == col_a.strip().lower():
                    ca = col
                if v == col_b.strip().lower():
                    cb = col
            if ca and cb:
                hidx = i
                break
        if not ca or not cb:
            sys.exit(f"ERROR: could not find {col_a!r} / {col_b!r} in first 25 rows.")
        for cells in rows[hidx + 1:]:
            pairs.append((cells.get(ca) or "", cells.get(cb) or ""))
    return pairs


# ---------------------------------------------------------------- ground truth --
def load_entities(path, sheet, id_col, count_col):
    counts, blanks, dupes = {}, 0, 0
    for cid_raw, cnt_raw in load_two_columns(path, sheet, id_col, count_col):
        cid = norm_id(cid_raw)
        if not cid:
            continue
        raw = str(cnt_raw).strip().replace(",", "")
        if raw == "":
            blanks += 1
            continue
        try:
            n = int(float(raw))
        except ValueError:
            blanks += 1
            continue
        if cid in counts:
            dupes += 1
            counts[cid] = max(counts[cid], n)
        else:
            counts[cid] = n
    return counts, blanks, dupes


def load_manual_status(path, sheet, id_col, status_col):
    out, unrec = {}, Counter()
    for cid_raw, st_raw in load_two_columns(path, sheet, id_col, status_col):
        cid = norm_id(cid_raw)
        s = str(st_raw or "").strip()
        if not cid or not s:
            continue
        sl = s.lower()
        if sl == "complete" or "further review" in sl or "51+" in sl:
            out[cid] = "resp"
        elif sl.startswith("no"):
            out[cid] = "nonresp"
        else:
            unrec[s] += 1
    return out, unrec


# --------------------------------------------------------------- the inventory --
def load_inventory(path):
    recs, has = {}, {}
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
        rd = csv.DictReader(fh)
        fields = {(c or "").strip() for c in (rd.fieldnames or [])}
        for col in ("s2_lane", "s2_nr", "s2_ran", "s2_is_bde", "s2_llm_tokens",
                    "s2_llm_responsiveness", "di_calls", "ocr_pages", "img_ocr_calls",
                    "img_ocr_qualifying", "img_ocr_ok", "img_decode_failed",
                    "bde_person_count", "nr_stage1", "elapsed_s", "programmatic"):
            has[col] = col in fields
        for r in rd:
            name = (r.get("file_name") or r.get("rel_path") or "").strip()
            cid = norm_id(name)
            if not cid or cid in recs:
                continue
            est      = _i(r.get("estimated_entities"))
            bpc      = _i(r.get("bde_person_count"))
            is_struct = _b(r.get("is_structured"))
            eff      = bpc if (is_struct and bpc > 0) else max(est, bpc)
            lane     = (r.get("suggested_lane") or "").strip().lower()
            s2lane   = (r.get("s2_lane") or "").strip().lower()
            # "programmatic" column added in 3.0.0; fall back to is_structured
            prog = _b(r.get("programmatic")) if has["programmatic"] else is_struct
            recs[cid] = dict(
                name=name, type=type_of(name), ext=(r.get("ext") or "").lower(),
                rel_path=(r.get("rel_path") or "").strip(),
                status=(r.get("status") or "").strip().lower(),
                searchable=_b(r.get("searchable")), is_struct=is_struct,
                programmatic=prog,
                est=est, bpc=bpc, eff=eff,
                bucket=(r.get("entity_bucket") or "").strip(),
                detail=(r.get("detail") or "").strip(),
                # stage 1
                lane=lane, is_bde=_b(r.get("is_bde")),
                nr1=_b(r.get("nr_stage1")) if has["nr_stage1"] else (lane == NR_LANE),
                llm_consulted=_b(r.get("llm_consulted")),
                tok1=_i(r.get("llm_tokens")),
                # stage 2
                s2_ran=_b(r.get("s2_ran")), s2_lane=s2lane, s2_nr=_b(r.get("s2_nr")),
                s2_is_bde=_b(r.get("s2_is_bde")), tok2=_i(r.get("s2_llm_tokens")),
                s2_level=(r.get("s2_llm_responsiveness") or "").strip(),
                s2_skip=(r.get("s2_skip_reason") or "").strip(),
                # OCR
                di=_i(r.get("di_calls")), ocr_pages=_i(r.get("ocr_pages")),
                img_q=_i(r.get("img_ocr_qualifying")), img_c=_i(r.get("img_ocr_calls")),
                img_ok=_i(r.get("img_ocr_ok")), img_fail=_i(r.get("img_decode_failed")),
                ocr_att=_b(r.get("ocr_attempted")), ocr_ok=_b(r.get("ocr")),
                elapsed=float(r.get("elapsed_s") or 0 or 0),
            )
    return recs, has


# ------------------------------------------------------------------- verdicts --
def stage1_call(a):
    if a["lane"] in RESP_LANES:
        return True
    if a["lane"] == NR_LANE:
        return False
    return None


def stage2_call(a):
    if not a["s2_ran"]:
        return None
    if a["s2_lane"] in RESP_LANES:
        return True
    if a["s2_lane"] == NR_LANE:
        return False
    return None


def pipeline_call(a):
    s1 = stage1_call(a)
    if s1 is None:
        return None
    if s1 is False:
        return False
    s2 = stage2_call(a)
    return True if s2 is None else s2


def union_call(a):
    s1, s2 = stage1_call(a), stage2_call(a)
    if s1 is None and s2 is None:
        return None
    return bool(s1) or bool(s2)


# -------------------------------------------------------------------- metrics --
def confusion(ids, recs, truth, pred):
    tp = fp = fn = tn = undet = 0
    fps, fns = [], []
    for cid in ids:
        p = pred(recs[cid])
        if p is None:
            undet += 1
            continue
        t = truth[cid]
        if t and p:       tp += 1
        elif p and not t: fp += 1; fps.append(cid)
        elif t and not p: fn += 1; fns.append(cid)
        else:             tn += 1
    return dict(TP=tp, FP=fp, FN=fn, TN=tn, undetermined=undet, fps=fps, fns=fns)


def stats(c):
    tp, fp, fn, tn = c["TP"], c["FP"], c["FN"], c["TN"]
    n = tp + fp + fn + tn
    return dict(
        N=n, TP=tp, FP=fp, FN=fn, TN=tn, undetermined=c["undetermined"],
        accuracy=pct(tp + tn, n), precision=pct(tp, tp + fp), recall=pct(tp, tp + fn),
        f1=pct(2 * tp, 2 * tp + fp + fn), specificity=pct(tn, tn + fp),
        nr_accuracy=pct(fn, fn + tn),
        r_accuracy=pct(fp, fp + tp),
        over_call=pct(fp, n), under_call=pct(fn, tp + fn),
        flagged_pct=pct(tp + fp, n),
    )


def fmt(x, nd=4):
    return "n/a" if x is None else f"{x:.{nd}f}"


def print_rnr(label, s, note=""):
    print("\n" + "=" * 74)
    print(f"NR/R -- {label}")
    if note:
        print(f"        {note}")
    print("=" * 74)
    print(f"  scored {s['N']:,}   (undetermined, excluded: {s['undetermined']:,})")
    print(f"  TP={s['TP']:,}  FP={s['FP']:,}  FN={s['FN']:,}  TN={s['TN']:,}")
    print(f"  recall      {fmt(s['recall'])}    <- misses: truly responsive, tool cleared")
    print(f"  precision   {fmt(s['precision'])}")
    print(f"  accuracy    {fmt(s['accuracy'])}    F1 {fmt(s['f1'])}")
    print(f"  NR accuracy {fmt(s['nr_accuracy'])}    of files CLEARED, share actually responsive"
          f"   (target < 0.05)")
    print(f"  R accuracy  {fmt(s['r_accuracy'])}    of files FLAGGED, share actually non-responsive"
          f" (target < 0.50)")
    print(f"  flagged     {fmt(s['flagged_pct'])}    over-call {fmt(s['over_call'])}"
          f"   under-call {fmt(s['under_call'])}")


def print_bde(label, s):
    print("\n" + "-" * 74)
    print(f"BDE -- {label}")
    print("-" * 74)
    print(f"  true BDEs {s['TP'] + s['FN']:,}   flagged {s['TP'] + s['FP']:,}   scored {s['N']:,}")
    print(f"  TP={s['TP']:,}  FP={s['FP']:,}  FN={s['FN']:,}  TN={s['TN']:,}")
    print(f"  recall {fmt(s['recall'])}   precision {fmt(s['precision'])}   "
          f"F1 {fmt(s['f1'])}   accuracy {fmt(s['accuracy'])}")
    print(f"  miss rate {fmt(s['under_call'])}  ({s['FN']:,} missed)")


# ============================================================ scaling helpers ==
def _read_manifest_root(inventory_path: str) -> str:
    try:
        with open(inventory_path + ".manifest.json", encoding="utf-8") as fh:
            return json.load(fh).get("root", "")
    except Exception:
        return ""


def _try_run_metrics():
    """Return scaling data dict, or None if unavailable."""
    try:
        from scaling_lib.metrics import run_metrics
        m = run_metrics()
    except ImportError:
        return None
    except Exception as exc:
        print(f"\n  (scaling-lib run_metrics() unavailable: {type(exc).__name__}: {exc})")
        return None
    completed = [t for t in m.tasks if t.status == "completed"]
    win_instances = {
        t.worker_instance for t in completed
        if t.worker_instance and
        any(t.file_name.lower().endswith(e) for e in _LEGACY_EXTS)
    }
    win_tasks = [t for t in completed if t.worker_instance in win_instances]
    lin_tasks = [t for t in completed if t.worker_instance not in win_instances]
    return {"m": m, "completed": completed,
            "win_instances": win_instances,
            "win_tasks": win_tasks, "lin_tasks": lin_tasks}


def _load_timing_snapshot(path: str):
    """Load a _timing.json written by collect_outputs.dump_timing.

    Returns the same dict shape as _try_run_metrics() so all downstream
    chart/sheet code can use either source without modification.
    """
    import types
    from datetime import datetime

    def _dt(s):
        if not s:
            return None
        try:
            return datetime.fromisoformat(s)
        except Exception:
            return None

    with open(path, encoding="utf-8") as fh:
        data = json.load(fh)

    def _task(t):
        return types.SimpleNamespace(
            file_name=       t.get("file_name", ""),
            status=          t.get("status", ""),
            worker_instance= t.get("worker_instance", ""),
            started_at=      _dt(t.get("started_at")),
            completed_at=    _dt(t.get("completed_at")),
            processing_s=    t.get("processing_s"),
            attempt_count=   t.get("attempt_count", 1),
            tokens_in=       t.get("tokens_in", 0),
            tokens_out=      t.get("tokens_out", 0),
            checkpoints=[
                types.SimpleNamespace(
                    label=      cp["label"],
                    duration_s= cp["duration_s"],
                    metadata=   cp.get("metadata", {}),
                )
                for cp in t.get("checkpoints", [])
            ],
        )

    tasks = [_task(t) for t in data.get("tasks", [])]
    m = types.SimpleNamespace(
        tasks=          tasks,
        total_tokens_in= data.get("total_tokens_in",  0),
        total_tokens_out=data.get("total_tokens_out", 0),
        files_completed= data.get("files_completed",  0),
        files_failed=    data.get("files_failed",     0),
        files_retried=   data.get("files_retried",    0),
        wall_clock_s=    data.get("wall_clock_s"),
        worker_count=    data.get("worker_count",     0),
        total_bytes=     data.get("total_bytes",      0),
    )
    completed = [t for t in tasks if t.status == "completed"]
    win_instances = {
        t.worker_instance for t in completed
        if t.worker_instance and
        any(t.file_name.lower().endswith(e) for e in _LEGACY_EXTS)
    }
    win_tasks = [t for t in completed if t.worker_instance in win_instances]
    lin_tasks = [t for t in completed if t.worker_instance not in win_instances]
    return {"m": m, "completed": completed,
            "win_instances": win_instances,
            "win_tasks": win_tasks, "lin_tasks": lin_tasks,
            "worker_config": data.get("worker_config")}


def _compute_worker_cost(tasks, win_instances, worker_config) -> dict | None:
    """Compute Linux worker-hours from actual per-worker task spans, then cost.

    Uses max(completed_at) - min(started_at) per distinct worker_instance so
    scale-in at the end of the run (one worker finishing last) is accounted for
    accurately — no naive replicas × wall_clock multiplication.
    """
    if not worker_config:
        return None

    by_worker: dict = defaultdict(lambda: {"starts": [], "ends": []})
    for t in tasks:
        if t.started_at and t.completed_at and t.worker_instance:
            by_worker[t.worker_instance]["starts"].append(t.started_at)
            by_worker[t.worker_instance]["ends"].append(t.completed_at)

    lin_hours = 0.0
    for instance, d in by_worker.items():
        if instance in win_instances:
            continue
        span_s = (max(d["ends"]) - min(d["starts"])).total_seconds()
        lin_hours += max(0.0, span_s) / 3600

    vcpu       = float(worker_config.get("vcpu", 0) or 0)
    gb         = float(worker_config.get("gb",   0) or 0)
    gpu        = int(worker_config.get("gpu",    0) or 0)
    price_vcpu = worker_config.get("price_vcpu_hr")
    price_gb   = worker_config.get("price_gb_hr")
    price_gpu  = worker_config.get("price_gpu_hr")

    vcpu_cost = (lin_hours * vcpu * price_vcpu) if price_vcpu is not None else None
    gb_cost   = (lin_hours * gb   * price_gb)   if price_gb   is not None else None
    gpu_cost  = (lin_hours * gpu  * price_gpu)  if (gpu and price_gpu is not None) else (
                 0.0 if gpu == 0 else None)
    priced    = [c for c in (vcpu_cost, gb_cost, gpu_cost) if c is not None]
    total     = sum(priced) if len(priced) == 3 or (gpu == 0 and len(priced) == 2) else None
    return {
        "lin_hours":      lin_hours,
        "vcpu":           vcpu,
        "gb":             gb,
        "gpu":            gpu,
        "profile_type":   worker_config.get("profile_type", ""),
        "workload_profile": worker_config.get("workload_profile", ""),
        "app_name":       worker_config.get("app_name", ""),
        "location":       worker_config.get("location", ""),
        "price_vcpu_hr":  price_vcpu,
        "price_gb_hr":    price_gb,
        "price_gpu_hr":   price_gpu,
        "vcpu_cost":      vcpu_cost,
        "gb_cost":        gb_cost,
        "gpu_cost":       gpu_cost,
        "total":          total,
    }


def _wall_tput(tasks) -> float | None:
    starts = [t.started_at   for t in tasks if t.started_at]
    ends   = [t.completed_at for t in tasks if t.completed_at]
    if not starts or not ends:
        return None
    span = (max(ends) - min(starts)).total_seconds()
    return len(tasks) / span * 3600 if span > 0 else None


def _tstats(values: list[float]) -> dict:
    n = len(values)
    if not n:
        return dict(count=0, total_s=0.0, avg_s=None, median_s=None,
                    p95_s=None, min_s=None, max_s=None)
    s = sorted(values)
    return dict(count=n, total_s=sum(s),
                avg_s=statistics.mean(s), median_s=statistics.median(s),
                p95_s=s[min(int(n * 0.95), n - 1)], min_s=s[0], max_s=s[-1])


def _group_cps(tasks) -> dict:
    totals: dict = defaultdict(lambda: {"count": 0, "total_s": 0.0,
                                         "tokens_in": 0, "tokens_out": 0})
    for task in tasks:
        for cp in task.checkpoints:
            totals[cp.label]["count"]     += 1
            totals[cp.label]["total_s"]   += cp.duration_s
            totals[cp.label]["tokens_in"] += int(cp.metadata.get("tokens_in",  0) or 0)
            totals[cp.label]["tokens_out"]+= int(cp.metadata.get("tokens_out", 0) or 0)
    return {
        lbl: {**v, "avg_s": v["total_s"] / v["count"] if v["count"] else None}
        for lbl, v in totals.items()
    }


def _fmt_hms(s: float | None) -> str:
    if s is None:
        return "n/a"
    h, rem = divmod(int(s), 3600)
    m, sec = divmod(rem, 60)
    return f"{h}:{m:02d}:{sec:02d}" if h else f"{m}:{sec:02d}"


# ================================================================ Excel writer ==

# -- colour palette (timing_report.py conventions) --
_C_TITLE    = "1F3864"   # dark navy   — title bars
_C_SECT     = "2E5496"   # medium blue — section bars
_C_KEY_BG   = "D9EAD3"   # light green — KPI rows
_C_HEAD     = "BDD7EE"   # light blue  — column headers
_C_ALT      = "EBF3FB"   # very light blue — alternating data rows
_C_WIN      = "FCE4D6"   # light orange — Windows worker
_C_LIN      = "E2EFDA"   # light green  — Linux worker
_C_COST     = "FFF2CC"   # light yellow — cost / money rows
_C_WHITE    = "FFFFFF"
_C_GTXT     = "808080"   # grey text (units)
_C_TP       = "C6EFCE"   # TP fill (green)
_C_FP       = "FFEB9C"   # FP fill (yellow)
_C_FN       = "FFC7CE"   # FN fill (red)
_C_TN       = "D9D9D9"   # TN fill (grey)
_C_TOTAL    = "FCE4D6"   # total/overall row (salmon)
_C_AGG      = "EDEDED"   # aggregate/subtotal rows
_C_NR_PASS  = "C6EFCE"   # NR accuracy < 5% → green
_C_NR_FAIL  = "FFC7CE"   # NR accuracy ≥ 5% → red
_C_MBT_HEAD = "D9E1F2"   # Metrics-by-Type header row

_CLS_FILL = {"TP": _C_TP, "FP": _C_FP, "FN": _C_FN, "TN": _C_TN}

_PCT = "0.00%"
_NUM = "#,##0"
_DEC = "0.00"
_USD = '$#,##0.00'
_USD4 = '$#,##0.0000'


def _xl(cell, bold=False, bg=None, fg="1F1F1F", align="left",
        num_fmt=None, size=11, wrap=False):
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(bold=bold, color=fg, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if num_fmt:
        cell.number_format = num_fmt


def _autofit(ws, mn=8, mx=60):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        best = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(best + 2, mn), mx)


def _title(ws, row, text, span, col_count=None):
    col_count = col_count or span
    ws.row_dimensions[row].height = 22
    c = ws.cell(row, 1, text)
    _xl(c, bold=True, bg=_C_TITLE, fg=_C_WHITE, size=13)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        for col in range(2, span + 1):
            _xl(ws.cell(row, col), bg=_C_TITLE)
    return row + 1


def _section(ws, row, text, span):
    ws.row_dimensions[row].height = 17
    c = ws.cell(row, 1, text)
    _xl(c, bold=True, bg=_C_SECT, fg=_C_WHITE)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        for col in range(2, span + 1):
            _xl(ws.cell(row, col), bg=_C_SECT)
    return row + 1


def _kpi(ws, row, label, value, unit="", val_fmt=None):
    ws.row_dimensions[row].height = 20
    _xl(ws.cell(row, 1, label), bold=True, size=12, bg=_C_KEY_BG)
    _xl(ws.cell(row, 2, value), bold=True, size=14, bg=_C_KEY_BG, align="right",
        num_fmt=val_fmt)
    _xl(ws.cell(row, 3, unit), size=10, fg=_C_GTXT, bg=_C_KEY_BG)
    return row + 1


def _kv(ws, row, key, val, val_bg=None, val_fmt=None, key_bg=None, span=3):
    _xl(ws.cell(row, 1, key), bold=True, bg=key_bg)
    _xl(ws.cell(row, 2, val), bg=val_bg, align="right", num_fmt=val_fmt)
    return row + 1


def _blank(ws, row, h=5):
    ws.row_dimensions[row].height = h
    return row + 1


def _hdrs(ws, row, cols, bg=None):
    bg = bg or _C_HEAD
    for c, h in enumerate(cols, 1):
        _xl(ws.cell(row, c, h), bold=True, bg=bg)
    return row + 1


def _fv(d, k):
    v = d.get(k)
    return round(v, 2) if v is not None else "n/a"


# ----------------------------------------------------------------- sheet: Run Info
def _sheet_run_info(wb, rp):
    ws = wb.active
    ws.title = "Run Info"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 50

    row = 1
    row = _title(ws, row, "pii_triage COMBINED-RUN SCORECARD", 3)
    row = _blank(ws, row)
    row = _section(ws, row, "SOURCES", 3)
    row = _kv(ws, row, "Inventory", rp["inventory_path"])
    row = _kv(ws, row, "Entities export", rp["entities_path"])
    row = _kv(ws, row, "Manual sheet", rp["manual_path"] or "(not provided)")
    row = _kv(ws, row, "Generated", rp["generated"])
    row = _blank(ws, row)

    row = _section(ws, row, "PARAMETERS", 3)
    row = _kv(ws, row, "Pipeline", rp["pipeline_desc"])
    absent_note = (f"AUTO → '{rp['mode']}' "
                   f"({'export has zero-count rows' if rp['mode'] == 'unreviewed' else 'no zero-count rows in export'})")
    row = _kv(ws, row, "Absent-means mode",
              absent_note if rp["absent_auto"] else f"{rp['mode']} (explicit)")
    row = _kv(ws, row, "BDE threshold",
              f"Total Entities > {rp['bde_threshold']} (i.e. ≥ {rp['bde_threshold'] + 1})")
    row = _blank(ws, row)

    row = _section(ws, row, "POPULATION", 3)
    row = _kpi(ws, row, "Files in inventory", rp["n_inventory"], "total files scanned",
               val_fmt=_NUM)
    row = _kpi(ws, row, "Scored", rp["n_scored"], "have a ground-truth verdict",
               val_fmt=_NUM)
    resp_pct = pct(rp["n_responsive"], rp["n_scored"])
    row = _kpi(ws, row, "Truly responsive", rp["n_responsive"],
               f"{resp_pct:.1%} of scored", val_fmt=_NUM)
    row = _kpi(ws, row, f"Truly BDE (count > {rp['bde_threshold']})",
               rp["n_bde"], "", val_fmt=_NUM)

    mr = rp["match_rate"]
    mr_bg = _C_NR_FAIL if mr < 0.95 else None
    row = _kv(ws, row, "Entity export ID match rate",
              f"{mr:.1%}" + (" ← LOW — check --id-col" if mr < 0.95 else ""),
              val_bg=mr_bg)
    row = _kv(ws, row, "In export, not in run", rp["reviewed_not_scanned"])
    if rp["mode"] == "unreviewed":
        row = _kv(ws, row, "In run, not reviewed", rp["n_not_reviewed"])

    if rp["manual_path"]:
        row = _blank(ws, row)
        row = _section(ws, row, "TRUTH CROSS-CHECK (entities vs manual status)", 3)
        row = _kv(ws, row, "Common files (both sources)", rp.get("xcheck_n", "n/a"))
        row = _kv(ws, row, "Agreement", rp.get("xcheck_agree_pct", "n/a"))
        if rp.get("xcheck_disagree"):
            for k, v in rp["xcheck_disagree"]:
                row = _kv(ws, row, f"  {k}", v)


# ----------------------------------------------------------------- sheet: Cost
def _sheet_cost(wb, rp, scaling):
    ws = wb.create_sheet("Cost")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40

    row = 1
    row = _title(ws, row, "RUN COST", 3)

    # OCR section
    row = _blank(ws, row)
    row = _section(ws, row, "DOCUMENT INTELLIGENCE (OCR)", 3)
    row = _kv(ws, row, "DI calls (total)", rp["di_calls"], val_fmt=_NUM)
    row = _kv(ws, row, "  Full-file OCR calls",
              sum(1 for r in rp["_recs"].values() if r["ocr_att"]), val_fmt=_NUM)
    row = _kv(ws, row, "  Embedded-image DI calls", rp["img_calls"], val_fmt=_NUM)
    row = _kv(ws, row, "Billable pages (total)", rp["di_pages"], val_fmt=_NUM)
    row = _kv(ws, row, "  Full-file pages",
              sum(r["ocr_pages"] for r in rp["_recs"].values()), val_fmt=_NUM)
    row = _kv(ws, row, "  Embedded-image pages", rp["img_calls"], val_fmt=_NUM)
    row = _kv(ws, row, "Qualifying embedded images", rp["img_q"], val_fmt=_NUM)
    row = _kv(ws, row, "Embedded images → returned text", rp["img_ok"], val_fmt=_NUM)
    if rp["img_fail"]:
        row = _kv(ws, row, "!! Images failed to decode (Pillow?)", rp["img_fail"],
                  val_bg=_C_NR_FAIL, val_fmt=_NUM)
    if rp["di_cost"] is not None:
        row = _kv(ws, row, "DI cost", rp["di_cost"],
                  val_bg=_C_COST, val_fmt=_USD)
    else:
        row = _kv(ws, row, "DI cost", "(pass --price-per-1k-pages to price this)")

    # LLM section
    row = _blank(ws, row)
    row = _section(ws, row, "LLM TOKENS", 3)
    row = _kv(ws, row, "Stage-1 tokens (inventory)",
              rp["tok1"], val_fmt=_NUM)
    _xl(ws.cell(row - 1, 3, f"{rp['n_llm1']:,} files consulted"), size=10, fg=_C_GTXT)
    row = _kv(ws, row, "Stage-2 tokens (inventory)",
              rp["tok2"], val_fmt=_NUM)
    _xl(ws.cell(row - 1, 3, f"{rp['n_llm2']:,} files"), size=10, fg=_C_GTXT)
    row = _kv(ws, row, "Total tokens (inventory)", rp["tok1"] + rp["tok2"],
              val_fmt=_NUM)

    if scaling:
        m = scaling["m"]
        row = _blank(ws, row)
        row = _kv(ws, row, "Tokens IN  (run_metrics — all attempts)",
                  m.total_tokens_in, val_bg=_C_COST, val_fmt=_NUM)
        row = _kv(ws, row, "Tokens OUT (run_metrics — all attempts)",
                  m.total_tokens_out, val_bg=_C_COST, val_fmt=_NUM)
        if rp["price_per_1k_in"] and rp["price_per_1k_out"]:
            exact_cost = (m.total_tokens_in  / 1000.0 * rp["price_per_1k_in"] +
                          m.total_tokens_out / 1000.0 * rp["price_per_1k_out"])
            row = _kv(ws, row, "LLM cost (exact, from run_metrics)",
                      exact_cost, val_bg=_C_COST, val_fmt=_USD4)
            _xl(ws.cell(row - 1, 3,
                        f"${rp['price_per_1k_in']}/1k in, ${rp['price_per_1k_out']}/1k out"),
                size=10, fg=_C_GTXT)
        else:
            row = _kv(ws, row, "LLM cost", "(pass --price-per-1k-in/out to price this)")
    else:
        lo, hi = rp["llm_lo"], rp["llm_hi"]
        if hi:
            row = _kv(ws, row, "LLM cost (low estimate)",  lo, val_bg=_C_COST, val_fmt=_USD4)
            row = _kv(ws, row, "LLM cost (high estimate)", hi, val_bg=_C_COST, val_fmt=_USD4)
            _xl(ws.cell(row - 1, 3, "in/out split unknown → range"), size=10, fg=_C_GTXT)
        else:
            row = _kv(ws, row, "LLM cost", "(pass --price-per-1k-in/out to price this)")

    # Compute (Workers)
    wc = rp.get("worker_cost")
    if wc:
        row = _blank(ws, row)
        row = _section(ws, row, "COMPUTE (LINUX WORKERS)", 3)
        row = _kv(ws, row, "Container App", wc["app_name"])
        row = _kv(ws, row, "Workload profile", wc["workload_profile"])
        row = _kv(ws, row, "Location", wc["location"])
        row = _kv(ws, row, "vCPU per replica", wc["vcpu"])
        row = _kv(ws, row, "GiB RAM per replica", wc["gb"])
        if wc["gpu"]:
            row = _kv(ws, row, "GPU per replica",
                      f"{wc['gpu']}× {wc['profile_type']}" if wc["profile_type"] else wc["gpu"])
        row = _kv(ws, row, "Actual worker-hours",
                  round(wc["lin_hours"], 2), val_fmt=_DEC)
        _xl(ws.cell(row - 1, 3,
                    "sum of per-worker spans (handles scale-in/out)"),
            size=10, fg=_C_GTXT)
        if wc["price_vcpu_hr"] is not None:
            row = _kv(ws, row, "Price per vCPU-hour", wc["price_vcpu_hr"], val_fmt=_USD4)
            row = _kv(ws, row, "Price per GiB-hour",  wc["price_gb_hr"],   val_fmt=_USD4)
            if wc["gpu"]:
                if wc["price_gpu_hr"] is not None:
                    row = _kv(ws, row, "Price per GPU-hour", wc["price_gpu_hr"], val_fmt=_USD4)
                else:
                    row = _kv(ws, row, "Price per GPU-hour",
                              "(not found in Retail Prices API — add manually)")
            _xl(ws.cell(row - 1, 3, "from Azure Retail Prices API"), size=10, fg=_C_GTXT)
        if wc["total"] is not None:
            row = _kv(ws, row, "vCPU cost", wc["vcpu_cost"], val_fmt=_USD)
            row = _kv(ws, row, "RAM cost",  wc["gb_cost"],   val_fmt=_USD)
            if wc["gpu"]:
                row = _kv(ws, row, "GPU cost", wc["gpu_cost"], val_fmt=_USD)
            row = _kv(ws, row, "Compute cost", wc["total"],
                      val_bg=_C_COST, val_fmt=_USD)
        else:
            row = _kv(ws, row, "Compute cost",
                      "(pricing unavailable — Retail Prices API may not cover this region/profile)")

    # Totals
    di_cost = rp["di_cost"]
    llm_hi  = rp["llm_hi"]
    if scaling and rp["price_per_1k_in"] and rp["price_per_1k_out"]:
        m = scaling["m"]
        llm_exact = (m.total_tokens_in  / 1000.0 * rp["price_per_1k_in"] +
                     m.total_tokens_out / 1000.0 * rp["price_per_1k_out"])
        llm_hi = llm_exact

    compute_total = wc["total"] if wc and wc.get("total") is not None else None
    have_totals   = sum(v is not None for v in [di_cost, llm_hi, compute_total]) >= 2

    if have_totals:
        row = _blank(ws, row)
        row = _section(ws, row, "TOTALS", 3)
        parts = {k: v for k, v in [("OCR",     di_cost),
                                     ("LLM",     llm_hi),
                                     ("Compute", compute_total)]
                 if v is not None}
        grand = sum(parts.values())
        row = _kv(ws, row, "Total cost (" + " + ".join(parts) + ")",
                  grand, val_bg=_C_COST, val_fmt=_USD)
        for label, cost in parts.items():
            row = _kv(ws, row, f"  {label} share", f"{pct(cost, grand):.1%}")

    # Single-pass saving
    row = _blank(ws, row)
    row = _section(ws, row, "SINGLE-PASS SAVING vs old Anna→Daniel two-pass workflow", 3)
    row = _kv(ws, row, "Stage-1 survivors (sent to stage 2 / reviewer)",
              rp["n_surv"], val_fmt=_NUM)
    _xl(ws.cell(row - 1, 3,
                f"{pct(rp['n_surv'], rp['n_inventory']):.1%} of all files"),
        size=10, fg=_C_GTXT)
    row = _kv(ws, row, "DI calls a 2nd pass would repeat", rp["dup_di"], val_fmt=_NUM)
    row = _kv(ws, row, "Pages a 2nd pass would repeat", rp["dup_pages"], val_fmt=_NUM)
    total_di = rp["di_calls"] + rp["dup_di"]
    row = _kv(ws, row, "DI calls saved",
              f"{pct(rp['dup_di'], total_di):.1%} ({rp['dup_di']:,} of {total_di:,})")
    if di_cost is not None and rp["di_pages"]:
        saved_cost = rp["dup_pages"] / 1000.0 * rp["price_per_1k_pages"]
        row = _kv(ws, row, "DI money saved", saved_cost, val_bg=_C_COST, val_fmt=_USD)


# ----------------------------------------------------------------- sheet: Timing
def _sheet_timing(wb, scaling):
    ws = wb.create_sheet("Timing")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 36

    row = 1
    row = _title(ws, row, "PROCESSING PERFORMANCE", 3)
    if scaling is None:
        row = _blank(ws, row)
        _xl(ws.cell(row, 1,
                    "scaling-lib run_metrics() not available — no timing data."),
            fg="808080")
        return

    m         = scaling["m"]
    completed = scaling["completed"]
    win_tasks = scaling["win_tasks"]
    lin_tasks = scaling["lin_tasks"]
    ref_tasks = lin_tasks if lin_tasks else completed
    # Use Linux-only times for per-file stats when Windows workers are also present,
    # since Windows tasks are conversion-only steps (COM call), not full file reviews.
    mixed     = bool(win_tasks and lin_tasks)
    all_times = [t.processing_s for t in ref_tasks if t.processing_s is not None]
    st        = _tstats(all_times)
    ref_tput  = _wall_tput(ref_tasks)
    all_tput  = _wall_tput(completed)
    tok_in    = m.total_tokens_in
    tok_out   = m.total_tokens_out
    time_sfx  = " (Linux)" if mixed else ""

    row = _blank(ws, row)
    row = _section(ws, row, "KEY METRICS", 3)
    if st["avg_s"] is not None:
        row = _kpi(ws, row, f"Avg time per file{time_sfx}",    round(st["avg_s"],    2), "seconds")
        row = _kpi(ws, row, f"Median time per file{time_sfx}", round(st["median_s"], 2), "seconds")
        row = _kpi(ws, row, f"P95 time per file{time_sfx}",    round(st["p95_s"],    2), "seconds")
    if ref_tput:
        row = _kpi(ws, row, "Files per hour",       round(ref_tput, 0),      "", val_fmt=_NUM)
        row = _kpi(ws, row, "Files per 24 hours",   round(ref_tput * 24, 0), "", val_fmt=_NUM)
        row = _kpi(ws, row, "Projected time for 200,000 files",
                   round(200_000 / ref_tput, 1), "hours")
    row = _blank(ws, row)
    row = _kpi(ws, row, "Tokens IN",  tok_in,  "", val_fmt=_NUM)
    row = _kpi(ws, row, "Tokens OUT", tok_out, "", val_fmt=_NUM)

    row = _blank(ws, row)
    row = _section(ws, row, "JOB OVERVIEW", 3)
    row = _kv(ws, row,
              "Tasks completed (Linux + Windows)" if mixed else "Files completed",
              m.files_completed, val_fmt=_NUM)
    if mixed:
        row = _kv(ws, row, "  Linux detections", len(lin_tasks), val_fmt=_NUM)
        row = _kv(ws, row, "  Windows conversions", len(win_tasks), val_fmt=_NUM)
    row = _kv(ws, row,
              "Tasks failed / dead-lettered" if mixed else "Files failed / dead-lettered",
              m.files_failed, val_fmt=_NUM)
    row = _kv(ws, row,
              "Tasks retried" if mixed else "Files retried",
              m.files_retried, val_fmt=_NUM)
    row = _kv(ws, row, "Total data",
              f"{m.total_bytes / 1_048_576:.1f} MB")
    row = _kv(ws, row, "Wall clock",  _fmt_hms(m.wall_clock_s))
    row = _kv(ws, row, "Worker instances", m.worker_count, val_fmt=_NUM)
    if scaling["win_instances"]:
        row = _kv(ws, row, "Windows worker(s)",
                  ", ".join(sorted(scaling["win_instances"])))
    if all_tput and ref_tput and abs(all_tput - ref_tput) > 1:
        row = _kv(ws, row, "Throughput all workers (tasks/hr)", round(all_tput, 0),
                  val_fmt=_NUM)

    cps = _group_cps(completed)
    if cps:
        row = _blank(ws, row)
        row = _section(ws, row, "STAGE TIMING (all workers)", 7)
        row = _hdrs(ws, row,
                    ["Stage", "Calls", "Avg (s)", "Total (s)",
                     "Tokens IN", "Tokens OUT", "Notes"])
        for lbl, v in sorted(cps.items(), key=lambda kv: -(kv[1]["total_s"])):
            avg = round(v["avg_s"], 3) if v["avg_s"] is not None else ""
            ti  = v["tokens_in"]  if lbl in ("azure_openai", "llm", "stage2") else ""
            to  = v["tokens_out"] if lbl in ("azure_openai", "llm", "stage2") else ""
            vals = [lbl, v["count"], avg, round(v["total_s"], 1), ti, to, ""]
            for c, val in enumerate(vals, 1):
                al = "right" if c > 1 and c < 7 else "left"
                _xl(ws.cell(row, c, val), align=al)
            row += 1

    # Windows vs Linux
    if win_tasks and lin_tasks:
        row = _blank(ws, row)
        row = _section(ws, row, "WINDOWS vs LINUX", 4)
        row = _hdrs(ws, row, ["Metric", "All Workers", "Windows", "Linux"])

        all_st = _tstats([t.processing_s for t in completed
                          if t.processing_s is not None])
        win_st = _tstats([t.processing_s for t in win_tasks
                          if t.processing_s is not None])
        lin_st = _tstats([t.processing_s for t in lin_tasks
                          if t.processing_s is not None])
        win_tp = _wall_tput(win_tasks)
        lin_tp = _wall_tput(lin_tasks)

        def _cmp(label, a, w, l):
            nonlocal row
            ws.cell(row, 1, label).font = __import__("openpyxl").styles.Font(bold=True)
            for col, (val, bg) in enumerate(
                    [(a, None), (w, _C_WIN), (l, _C_LIN)], 2):
                _xl(ws.cell(row, col, val), bg=bg, align="right")
            row += 1

        _cmp("Task count",
             all_st["count"], win_st["count"], lin_st["count"])
        for lbl2, key in [("Avg / task (s)", "avg_s"), ("Median (s)", "median_s"),
                           ("P95 (s)", "p95_s")]:
            _cmp(lbl2, _fv(all_st, key), _fv(win_st, key), _fv(lin_st, key))
        _cmp("Throughput (tasks/hr)",
             round(all_tput, 0) if all_tput else "n/a",
             round(win_tp,   0) if win_tp   else "n/a",
             round(lin_tp,   0) if lin_tp   else "n/a")
        _cmp("Tasks / 24 hrs",
             round(all_tput * 24, 0) if all_tput else "n/a",
             round(win_tp   * 24, 0) if win_tp   else "n/a",
             round(lin_tp   * 24, 0) if lin_tp   else "n/a")

        lin_times = [t.processing_s for t in lin_tasks if t.processing_s is not None]
        if lin_times:
            lin_avg  = statistics.mean(lin_times)
            n_lin    = len({t.worker_instance for t in lin_tasks if t.worker_instance})
            proj_tput = n_lin * 3600 / lin_avg
            n_total  = len(lin_tasks)   # unique files (Windows tasks are conversion steps only)
            row = _blank(ws, row)
            row = _section(ws, row, "PROJECTION: all files on Linux workers only", 4)
            row = _kv(ws, row, "Linux avg / file (s)", round(lin_avg, 2))
            row = _kv(ws, row, "Linux workers", n_lin, val_fmt=_NUM)
            row = _kv(ws, row, "Projected throughput (files/hr)", round(proj_tput, 0),
                      val_fmt=_NUM)
            row = _kv(ws, row, "Projected total time (hr)",
                      round(n_total / proj_tput, 2))
            if all_tput:
                row = _kv(ws, row, "Speed-up vs actual",
                          round(proj_tput / all_tput, 2))

    _autofit(ws, mn=10, mx=50)


# ----------------------------------------------------------------- timeline data prep
def _build_timeline_data(scaling):
    """Bucket per-minute: file completions by worker type, tokens, rate-limit events."""
    m   = scaling["m"]
    win = scaling["win_instances"]

    completed = [t for t in m.tasks
                 if t.status == "completed" and t.completed_at and t.started_at]
    if not completed:
        return None

    job_start = min(t.started_at   for t in completed)
    job_end   = max(t.completed_at for t in completed)
    n_mins    = max(2, int((job_end - job_start).total_seconds() / 60) + 2)

    lin_files = [0] * n_mins
    win_files = [0] * n_mins
    tok_in    = [0] * n_mins
    tok_out   = [0] * n_mins

    for t in completed:
        b = min(n_mins - 1, max(0,
                int((t.completed_at - job_start).total_seconds() / 60)))
        (win_files if t.worker_instance in win else lin_files)[b] += 1
        tok_in[b]  += t.tokens_in
        tok_out[b] += t.tokens_out

    # Rate-limit detection: two sources.
    # Reactive (429): azure_openai checkpoints that are duration outliers — the
    #   backoff sleep is inside that checkpoint, making it anomalously long.
    # Proactive (headroom): azure_openai_rate_limit_wait checkpoints with
    #   meaningful duration — the client slept before calling to avoid hitting
    #   the limit.  These are entirely missed by the duration-outlier heuristic.
    aoai_durs = sorted(
        cp.duration_s
        for t in m.tasks
        for cp in t.checkpoints
        if cp.label == "azure_openai" and cp.duration_s > 0
    )
    rl_threshold = None
    rl_events    = [0] * n_mins
    if len(aoai_durs) >= 3:
        med          = aoai_durs[len(aoai_durs) // 2]
        rl_threshold = max(30.0, 3.0 * med)

    for t in m.tasks:
        if not t.started_at:
            continue
        offset = 0.0
        for cp in t.checkpoints:
            is_reactive   = (cp.label == "azure_openai"
                             and rl_threshold is not None
                             and cp.duration_s >= rl_threshold)
            is_preemptive = (cp.label == "azure_openai_rate_limit_wait"
                             and cp.duration_s > 1.0)
            if is_reactive or is_preemptive:
                try:
                    elapsed = (t.started_at + datetime.timedelta(seconds=offset)
                               - job_start).total_seconds()
                except TypeError:
                    elapsed = (t.started_at.replace(tzinfo=None)
                               + datetime.timedelta(seconds=offset)
                               - job_start.replace(tzinfo=None)).total_seconds()
                b = min(n_mins - 1, max(0, int(elapsed / 60)))
                rl_events[b] += 1
            offset += cp.duration_s

    # Cumulative series (simple running sum)
    cum_lin = [0] * n_mins
    cum_win = [0] * n_mins
    cum_in  = [0] * n_mins
    cum_out = [0] * n_mins
    for i in range(n_mins):
        p = i - 1
        cum_lin[i] = (cum_lin[p] if i else 0) + lin_files[i]
        cum_win[i] = (cum_win[p] if i else 0) + win_files[i]
        cum_in[i]  = (cum_in[p]  if i else 0) + tok_in[i]
        cum_out[i] = (cum_out[p] if i else 0) + tok_out[i]

    return dict(
        n_mins       = n_mins,
        job_start    = job_start,
        lin_files    = lin_files,
        win_files    = win_files,
        tok_in       = tok_in,
        tok_out      = tok_out,
        rl_events    = rl_events,
        cum_lin      = cum_lin,
        cum_win      = cum_win,
        cum_in       = cum_in,
        cum_out      = cum_out,
        has_win      = bool(win and any(win_files)),
        has_tokens   = bool(any(tok_in) or any(tok_out)),
        has_rl       = bool(any(rl_events)),
        rl_threshold = rl_threshold,
        rl_count     = sum(rl_events),
        aoai_median  = aoai_durs[len(aoai_durs) // 2] if aoai_durs else None,
    )


# ----------------------------------------------------------------- sheet: Run Timeline
def _sheet_timeline(wb, scaling):
    if scaling is None:
        return
    td = _build_timeline_data(scaling)
    if td is None:
        return

    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter as gcl

    ws = wb.create_sheet("Run Timeline")
    n  = td["n_mins"]

    # ── Data table ───────────────────────────────────────────────────────────
    COLS = [
        "Minute",
        "Linux files/min",    "Windows conv./min",
        "Cum Linux files",    "Cum Windows conv.",
        "Tokens IN/min",      "Tokens OUT/min",
        "Cum Tokens IN",      "Cum Tokens OUT",
        "Rate-limit events",
    ]
    _hdrs(ws, 1, COLS)

    for i in range(n):
        r = i + 2
        ws.cell(r, 1,  i)
        ws.cell(r, 2,  td["lin_files"][i])
        ws.cell(r, 3,  td["win_files"][i])
        ws.cell(r, 4,  td["cum_lin"][i])
        ws.cell(r, 5,  td["cum_win"][i])
        ws.cell(r, 6,  td["tok_in"][i])
        ws.cell(r, 7,  td["tok_out"][i])
        ws.cell(r, 8,  td["cum_in"][i])
        ws.cell(r, 9,  td["cum_out"][i])
        ws.cell(r, 10, td["rl_events"][i])

    data_end = n + 1   # last data row (header is row 1, data starts row 2)

    for c in range(1, 11):
        ws.column_dimensions[gcl(c)].width = 18

    # ── Chart factory helpers ────────────────────────────────────────────────
    _SER_COLORS = ["4472C4", "ED7D31", "70AD47", "FFC000", "FF0000"]

    def _lc(title, y_label, col_list, anchor):
        """Line chart over one or more data columns (each with its own series)."""
        chart = LineChart()
        chart.title        = title
        chart.y_axis.title         = y_label
        chart.y_axis.delete        = False
        chart.y_axis.numFmt        = "#,##0"
        chart.y_axis.tickLblPos    = "nextTo"
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.title         = "Minutes from job start"
        chart.x_axis.delete        = False
        chart.x_axis.numFmt        = "0"
        chart.x_axis.tickLblPos    = "nextTo"
        chart.x_axis.majorTickMark = "out"
        chart.width        = 22   # cm
        chart.height       = 14   # cm

        cats = Reference(ws, min_col=1, min_row=2, max_row=data_end)
        for col in col_list:
            # min_row=1 so openpyxl reads the header cell as the series title
            data = Reference(ws, min_col=col, min_row=1, max_row=data_end)
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        for si, s in enumerate(chart.series):
            s.smooth = False
            s.graphicalProperties.line.solidFill = _SER_COLORS[si % len(_SER_COLORS)]
            s.graphicalProperties.line.width = 20000   # 2pt

        ws.add_chart(chart, anchor)
        return chart

    # Chart row starts 4 rows below the data table
    cr = data_end + 4

    # ── Chart 1: Files per minute ────────────────────────────────────────────
    file_cols = [2] + ([3] if td["has_win"] else [])
    chart1_title = ("Tasks Per Minute  (Linux files + Windows conversions)"
                    if td["has_win"] else "Files Processed Per Minute")
    _lc(chart1_title, "Tasks / minute" if td["has_win"] else "Files / minute",
        file_cols, f"A{cr}")

    # ── Chart 2: Cumulative files ────────────────────────────────────────────
    cum_cols = [4] + ([5] if td["has_win"] else [])
    chart2_title = ("Cumulative Tasks  (Linux files + Windows conversions)"
                    if td["has_win"] else "Cumulative Files Completed")
    _lc(chart2_title, "Tasks" if td["has_win"] else "Files", cum_cols, f"N{cr}")

    # ── Chart 3: Token rate per minute + RL events on secondary Y-axis ───────
    if td["has_tokens"]:
        from openpyxl.chart.marker import Marker as _Marker
        from openpyxl.chart.legend import Legend as _Legend

        chart3 = LineChart()
        chart3.title        = (
            "Token Rate Per Minute" +
            (f"  —  {td['rl_count']} rate-limit event(s)  (right axis)" if td["has_rl"] else "")
        )
        chart3.y_axis.title         = "Tokens / minute"
        chart3.y_axis.delete        = False
        chart3.y_axis.numFmt        = "#,##0"
        chart3.y_axis.tickLblPos    = "nextTo"
        chart3.y_axis.majorTickMark = "out"
        chart3.x_axis.title         = "Minutes from job start"
        chart3.x_axis.delete        = False
        chart3.x_axis.numFmt        = "0"
        chart3.x_axis.tickLblPos    = "nextTo"
        chart3.x_axis.majorTickMark = "out"
        chart3.width  = 22
        chart3.height = 17   # extra height so legend at bottom doesn't crowd lines
        # Do NOT set axIds on chart3 — keep openpyxl defaults (x=100, y=200).
        # Overriding them swaps crossAx references and corrupts the chart XML.

        # Legend below the plot area so it doesn't overlap the lines or axis labels
        _leg = _Legend()
        _leg.position = "b"
        _leg.overlay  = False
        chart3.legend = _leg

        cats = Reference(ws, min_col=1, min_row=2, max_row=data_end)
        for col in (6, 7):
            data = Reference(ws, min_col=col, min_row=1, max_row=data_end)
            chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(cats)
        for s, color in zip(chart3.series, ["4472C4", "70AD47"]):
            s.smooth = False
            s.graphicalProperties.line.solidFill = color
            s.graphicalProperties.line.width     = 20000

        if td["has_rl"]:
            # openpyxl's actual default axIds: catAx=10, valAx=100 (not 100/200).
            # The secondary lineChart must reference catAx=10 as its x-axis so its
            # series aligns on x, and crossAx=10 so the secondary valAx crosses the
            # category axis rather than the primary valAx.
            chart3b = LineChart()
            chart3b.y_axis.axId           = 300
            chart3b.y_axis.title          = "Rate-limit events"
            chart3b.y_axis.delete         = False
            chart3b.y_axis.numFmt         = "0"
            chart3b.y_axis.tickLblPos     = "high"   # labels on right-hand side
            chart3b.y_axis.majorTickMark  = "out"    # needed or Excel suppresses labels
            chart3b.y_axis.crosses        = "max"    # axis appears on right
            chart3b.y_axis.crossAx        = 10       # cross primary catAx (axId=10)
            chart3b.y_axis.axPos          = "r"      # RIGHT side — default is "l", which overlaps primary
            chart3b.y_axis.majorGridlines = None     # no extra gridlines over primary
            chart3b.y_axis.majorUnit      = 1        # integer ticks only
            chart3b.y_axis.scaling.min    = 0
            chart3b.y_axis.scaling.max    = max(max(td["rl_events"]), 1) + 1
            # Share primary catAx (actual default axId=10).
            # delete=True suppresses the duplicate catAx XML; Excel uses chart3's.
            chart3b.x_axis.axId           = 10
            chart3b.x_axis.delete         = True

            rl_data = Reference(ws, min_col=10, min_row=1, max_row=data_end)
            chart3b.add_data(rl_data, titles_from_data=True)
            chart3b.set_categories(cats)
            if chart3b.series:
                rl_s = chart3b.series[0]
                rl_s.smooth = False
                rl_s.graphicalProperties.line.solidFill = "C00000"
                rl_s.graphicalProperties.line.width     = 15000
                mk = _Marker()
                mk.symbol = "diamond"
                mk.size   = 8
                mk.graphicalProperties.solidFill      = "C00000"
                mk.graphicalProperties.line.solidFill = "C00000"
                rl_s.marker = mk

            chart3 += chart3b

        ws.add_chart(chart3, f"A{cr + 28}")

    # ── Chart 4: Cumulative tokens ───────────────────────────────────────────
    if td["has_tokens"]:
        _lc("Cumulative Tokens Consumed", "Tokens", [8, 9], f"N{cr + 28}")

    # ── Footnote ─────────────────────────────────────────────────────────────
    note_row = cr + (60 if td["has_tokens"] else 30)
    _xl(ws.cell(note_row, 1,
                "Minute 0 = first task started_at. "
                "Token counts are per-task totals from azure_openai checkpoints, "
                "placed at each task's completed_at."
                + (f" Rate-limit events: preemptive (azure_openai_rate_limit_wait > 1s) "
                   f"+ reactive 429s (azure_openai duration ≥ "
                   f"{td['rl_threshold']:.0f}s = 3× median). "
                   f"Minute estimated from checkpoint order within each task."
                   if td["has_rl"] else "")),
        fg=_C_GTXT, size=10, wrap=True)
    ws.row_dimensions[note_row].height = 36


# ----------------------------------------------------------------- sheet: NR/R Accuracy
def _sheet_nrr(wb, rp):
    ws = wb.create_sheet("NR-R Accuracy")
    NCOLS = 15
    row = 1
    row = _title(ws, row, "NR / R ACCURACY", NCOLS)
    row = _blank(ws, row)
    note = ws.cell(row, 1,
                   "Positive class = RESPONSIVE. "
                   "NR accuracy = FN/(FN+TN) = inaccurate-clear rate. "
                   "Target < 5%.")
    _xl(note, fg=_C_GTXT, size=10)
    row += 1
    row = _blank(ws, row)

    COLS = ["View", "Population", "N", "TP", "FP", "FN", "TN", "Undet",
            "Accuracy", "Recall", "Precision", "F1",
            "NR Accuracy\n(target<5%)", "R Accuracy", "Flagged%"]
    row = _hdrs(ws, row, COLS)
    ws.freeze_panes = f"A{row}"

    views = rp["nr_results"]   # {label: stats_dict}
    pops  = rp["nr_pops"]      # {label: population note}

    for lbl, s in views.items():
        r = row
        ws.cell(r, 1, lbl)
        ws.cell(r, 2, pops.get(lbl, ""))
        tp, fp, fn, tn, ud = s["TP"], s["FP"], s["FN"], s["TN"], s["undetermined"]
        # N as formula over the four counts
        ws.cell(r, 3, f"=D{r}+E{r}+F{r}+G{r}").number_format = _NUM
        ws.cell(r, 4, tp).number_format = _NUM
        ws.cell(r, 5, fp).number_format = _NUM
        ws.cell(r, 6, fn).number_format = _NUM
        ws.cell(r, 7, tn).number_format = _NUM
        ws.cell(r, 8, ud).number_format = _NUM
        # Derived metrics as formulas (stay correct if pasted as values into another workbook)
        for col, formula in [
            (9,  f'=IFERROR((D{r}+G{r})/C{r},"")'),   # Accuracy
            (10, f'=IFERROR(D{r}/(D{r}+F{r}),"")'),   # Recall
            (11, f'=IFERROR(D{r}/(D{r}+E{r}),"")'),   # Precision
            (12, f'=IFERROR(2*D{r}/(2*D{r}+E{r}+F{r}),"")'),  # F1
            (13, f'=IFERROR(F{r}/(F{r}+G{r}),"")'),   # NR Accuracy
            (14, f'=IFERROR(E{r}/(D{r}+E{r}),"")'),   # R Accuracy
            (15, f'=IFERROR((D{r}+E{r})/C{r},"")'),   # Flagged%
        ]:
            ws.cell(r, col, formula).number_format = _PCT

        # Colour the NR Accuracy cell based on the Python-computed value
        nr_bg = _C_NR_PASS if s["nr_accuracy"] < 0.05 else _C_NR_FAIL
        _xl(ws.cell(r, 13), bg=nr_bg, num_fmt=_PCT)

        row += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 38
    for col in "CDEFGHIJKLMNO":
        ws.column_dimensions[col].width = 12
    ws.row_dimensions[row - len(views)].height = 30  # header wrap


# ----------------------------------------------------------------- sheet: BDE Accuracy
def _sheet_bde(wb, rp):
    ws = wb.create_sheet("BDE Accuracy")
    NCOLS = 11
    row = 1
    row = _title(ws, row, f"BDE ACCURACY  (truth: Total Entities > {rp['bde_threshold']})",
                 NCOLS)
    row = _blank(ws, row)
    note = ws.cell(row, 1,
                   f"Off-by-one note: truth is count > {rp['bde_threshold']} "
                   f"(≥ {rp['bde_threshold']+1}); scanner flags is_bde when "
                   f"its count ≥ bde_threshold. Run scan with "
                   f"--bde-threshold {rp['bde_threshold']+1} to align exactly.")
    _xl(note, fg=_C_GTXT, size=10, wrap=True)
    ws.row_dimensions[row].height = 28
    row += 1
    row = _blank(ws, row)

    COLS = ["Definition", "N", "TP", "FP", "FN", "TN",
            "Recall", "Precision", "F1", "Accuracy", "Miss Rate"]
    row = _hdrs(ws, row, COLS)
    ws.freeze_panes = f"A{row}"

    for lbl, (s, _c) in rp["bde_results"].items():
        r = row
        ws.cell(r, 1, lbl)
        ws.cell(r, 2, f"=C{r}+D{r}+E{r}+F{r}").number_format = _NUM
        ws.cell(r, 3, s["TP"]).number_format = _NUM
        ws.cell(r, 4, s["FP"]).number_format = _NUM
        ws.cell(r, 5, s["FN"]).number_format = _NUM
        ws.cell(r, 6, s["TN"]).number_format = _NUM
        for col, formula in [
            (7,  f'=IFERROR(C{r}/(C{r}+E{r}),"")'),         # Recall
            (8,  f'=IFERROR(C{r}/(C{r}+D{r}),"")'),         # Precision
            (9,  f'=IFERROR(2*C{r}/(2*C{r}+D{r}+E{r}),"")'),# F1
            (10, f'=IFERROR((C{r}+F{r})/B{r},"")'),         # Accuracy
            (11, f'=IFERROR(E{r}/(C{r}+E{r}),"")'),         # Miss Rate
        ]:
            ws.cell(r, col, formula).number_format = _PCT
        row += 1

    # Count accuracy section
    ca = rp.get("count_acc")
    if ca:
        row = _blank(ws, row)
        row = _section(ws, row,
                       f"COUNT ACCURACY on {ca['n_true_bde']:,} true BDEs",
                       NCOLS)
        thr = ca["threshold"]
        row = _kv(ws, row, f"Raw estimate ≥ {thr}",
                  f"{ca['raw_ge']:,}  ({pct(ca['raw_ge'], ca['n_true_bde']):.1%})")
        row = _kv(ws, row, f"Effective ≥ {thr}  (incl. bde_person_count)",
                  f"{ca['eff_ge']:,}  ({pct(ca['eff_ge'], ca['n_true_bde']):.1%})")
        rescued = ca["rescued"]
        row = _kv(ws, row,
                  "Rescued by bde_person_count" if rescued >= 0 else
                  "Net change from bde_person_count (negative = counter corrected down)",
                  rescued)
        row = _kv(ws, row, "Still under-counted",
                  f"{ca['still_missed']:,}  "
                  f"({pct(ca['still_missed'], ca['n_true_bde']):.1%})")
        row = _kv(ws, row, "Mean |tool − human|", round(ca["mean_err"], 1))
        row = _kv(ws, row, "Median |tool − human|", ca["median_err"])

        if ca["under_by_type"]:
            row = _blank(ws, row)
            row = _section(ws, row, "Still-missed BDEs: breakdown", NCOLS)
            row = _hdrs(ws, row, ["Dimension", "Bucket", "Count"])
            for k, v in ca["under_by_type"].most_common():
                _xl(ws.cell(row, 1, "File type")); ws.cell(row, 2, k)
                _xl(ws.cell(row, 3, v), align="right", num_fmt=_NUM)
                row += 1
            for k, v in ca["under_by_searchable"].most_common():
                _xl(ws.cell(row, 1, "Searchable")); ws.cell(row, 2, k)
                _xl(ws.cell(row, 3, v), align="right", num_fmt=_NUM)
                row += 1

    ws.column_dimensions["A"].width = 46
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 13


# ----------------------------------------------------------------- sheet: Metrics by Type
# Three blocks side by side: Stage 1 | Stage 2 (graded only) | Pipeline (combined).
# When stage 2 did not run only Stage 1 is shown.
# Each block mirrors the HWE tracking template layout (12 cols, pasteable into tracker).
def _sheet_mbt(wb, rp):
    ws = wb.create_sheet("Metrics by Type")

    import openpyxl
    from openpyxl.utils import get_column_letter as gcl

    s2p = rp.get("stage2_present", False)

    HDRS = ["\xa0", "N", "TP", "FP", "FN", "TN",
            "Accuracy", "Recall", "Precision", "Under-call", "NR accuracy", "R accuracy"]

    title_fill  = openpyxl.styles.PatternFill("solid", fgColor=_C_SECT)
    header_fill = openpyxl.styles.PatternFill("solid", fgColor=_C_MBT_HEAD)
    total_fill  = openpyxl.styles.PatternFill("solid", fgColor=_C_TOTAL)
    agg_fill    = openpyxl.styles.PatternFill("solid", fgColor=_C_AGG)
    bold_font   = openpyxl.styles.Font(bold=True)
    title_font  = openpyxl.styles.Font(bold=True, color=_C_WHITE)
    hdr_font    = openpyxl.styles.Font(bold=True)

    # (display label, data-key prefix, 1-based starting column)
    BLOCKS = [("Stage 1", "s1", 1)]
    if s2p:
        BLOCKS.append(("Stage 2  (graded files only)", "s2", 14))
        BLOCKS.append(("Pipeline  (combined verdict)", "pipe", 27))

    def _wt(row, section_name, col0):
        c = ws.cell(row, col0, section_name)
        c.font = title_font; c.fill = title_fill
        ws.merge_cells(start_row=row, start_column=col0,
                       end_row=row, end_column=col0 + 11)
        for col in range(col0 + 1, col0 + 12):
            ws.cell(row, col).fill = title_fill

    def _wh(row, col0):
        for i, h in enumerate(HDRS):
            c = ws.cell(row, col0 + i, h)
            c.font = hdr_font; c.fill = header_fill

    def _wd(row, label, bucket, col0, fill=None, bold=False):
        tp, fp, fn, tn = bucket[0], bucket[1], bucket[2], bucket[3]
        NL  = gcl(col0 + 1); TPL = gcl(col0 + 2)
        FPL = gcl(col0 + 3); FNL = gcl(col0 + 4); TNL = gcl(col0 + 5)
        ws.cell(row, col0, label)
        ws.cell(row, col0 + 2, tp); ws.cell(row, col0 + 3, fp)
        ws.cell(row, col0 + 4, fn); ws.cell(row, col0 + 5, tn)
        ws.cell(row, col0 + 1,
                f"={TPL}{row}+{FPL}{row}+{FNL}{row}+{TNL}{row}").number_format = _NUM
        for i, formula in enumerate([
            f'=IFERROR(({TPL}{row}+{TNL}{row})/{NL}{row},"")',
            f'=IFERROR({TPL}{row}/({TPL}{row}+{FNL}{row}),"")',
            f'=IFERROR({TPL}{row}/({TPL}{row}+{FPL}{row}),"")',
            f'=IFERROR({FNL}{row}/({TPL}{row}+{FNL}{row}),"")',
            f'=IFERROR({FNL}{row}/({FNL}{row}+{TNL}{row}),"")',
            f'=IFERROR({FPL}{row}/({TPL}{row}+{FPL}{row}),"")',
        ]):
            ws.cell(row, col0 + 6 + i, formula).number_format = _PCT
        if fill or bold:
            for col in range(col0, col0 + 12):
                c = ws.cell(row, col)
                if fill: c.fill = fill
                if bold: c.font = bold_font

    # Write every block at the same row simultaneously so blocks stay in sync
    def _at(row, section_name):
        for title, _key, col0 in BLOCKS:
            _wt(row, f"{section_name} — {title}", col0)

    def _ah(row):
        for _t, _key, col0 in BLOCKS:
            _wh(row, col0)

    def _ad(row, label, bucket_fn, fill=None, bold=False):
        for _t, key, col0 in BLOCKS:
            _wd(row, label, bucket_fn(key), col0, fill=fill, bold=bold)

    row = 1

    # ── Section 1: By file type ──────────────────────────────────────────────
    _at(row, "BY FILE TYPE"); row += 1
    _ah(row); row += 1
    for ft in _FTYPE_ORDER:
        _ad(row, ft, lambda key, ft=ft: rp[f"mbt_{key}_ftype"][ft])
        row += 1
    _ad(row, "TOTAL / OVERALL",
        lambda key: rp[f"mbt_{key}_overall"], fill=total_fill, bold=True)
    row += 2

    # ── Section 2: By searchable ─────────────────────────────────────────────
    _at(row, "BY SEARCHABLE"); row += 1
    _ah(row); row += 1
    _ad(row, "searchable",
        lambda key: rp[f"mbt_{key}_search"]["searchable"],
        fill=agg_fill, bold=True); row += 1
    for ft in _FTYPE_ORDER:
        _ad(row, ft,
            lambda key, ft=ft: rp[f"mbt_{key}_sftype"]["searchable"][ft])
        row += 1
    _ad(row, "non-searchable",
        lambda key: rp[f"mbt_{key}_search"]["non-searchable"],
        fill=agg_fill, bold=True); row += 1
    for ft in _FTYPE_ORDER:
        _ad(row, ft,
            lambda key, ft=ft: rp[f"mbt_{key}_sftype"]["non-searchable"][ft])
        row += 1
    _ad(row, "TOTAL / OVERALL",
        lambda key: rp[f"mbt_{key}_overall"], fill=total_fill, bold=True)
    row += 2

    # ── Section 3: By searchable × structured ────────────────────────────────
    _at(row, "BY SEARCHABLE × STRUCTURED"); row += 1
    _ah(row); row += 1
    for (skey, pkey), show_ftypes in [
        (("searchable",     "structured"),     True),
        (("searchable",     "non-structured"), True),
        (("non-searchable", "structured"),     False),
        (("non-searchable", "non-structured"), False),
    ]:
        _ad(row, f"{skey} — {pkey}",
            lambda key, sk=skey, pk=pkey: rp[f"mbt_{key}_sp"][(sk, pk)],
            fill=agg_fill, bold=True); row += 1
        if show_ftypes:
            for ft in _FTYPE_ORDER:
                _ad(row, ft,
                    lambda key, sk=skey, pk=pkey, ft=ft:
                        rp[f"mbt_{key}_spftype"][(sk, pk)][ft])
                row += 1

    # Column widths — label col wide, count/metric cols narrow
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    for c in range(3, 13):
        ws.column_dimensions[gcl(c)].width = 12
    if s2p:
        ws.column_dimensions[gcl(13)].width = 4   # gap between blocks
        ws.column_dimensions[gcl(14)].width = 30
        ws.column_dimensions[gcl(15)].width = 10
        for c in range(16, 26):
            ws.column_dimensions[gcl(c)].width = 12
        ws.column_dimensions[gcl(26)].width = 4   # gap between blocks
        ws.column_dimensions[gcl(27)].width = 30
        ws.column_dimensions[gcl(28)].width = 10
        for c in range(29, 39):
            ws.column_dimensions[gcl(c)].width = 12


# ----------------------------------------------------------------- sheet: Stage 2 Detail
def _sheet_s2(wb, rp):
    if not rp.get("s2_detail"):
        return
    s2d = rp["s2_detail"]
    ws = wb.create_sheet("Stage 2 Detail")
    row = 1
    row = _title(ws, row, "STAGE 2 DETAIL", 5)

    _S2_LEVEL_DESC = {
        "clear_yes":  "LLM found PII that unambiguously identifies a person in scope",
        "likely_yes": "LLM found PII that probably identifies a person in scope, with minor uncertainty",
        "borderline": "LLM found signals that may or may not indicate in-scope PII — too close to call",
        "likely_no":  "LLM found minimal or ambiguous signals; probably no in-scope PII",
        "clear_no":   "LLM found no indicators of in-scope PII",
        "(none)":     "Stage 2 ran but returned no level (API error or unexpected response format)",
    }

    row = _blank(ws, row)
    row = _section(ws, row, "LEVEL DISTRIBUTION (graded files)", 5)
    row = _hdrs(ws, row, ["Level", "N", "Truly Responsive", "Truly Resp %", "Description"])
    for lv in s2d["levels"]:
        ws.cell(row, 1, lv["level"])
        _xl(ws.cell(row, 2, lv["n"]),         align="right", num_fmt=_NUM)
        _xl(ws.cell(row, 3, lv["true_resp"]), align="right", num_fmt=_NUM)
        _xl(ws.cell(row, 4, f"{lv['true_resp_pct']:.1%}"), align="right")
        _xl(ws.cell(row, 5, _S2_LEVEL_DESC.get(lv["level"], "")), fg=_C_GTXT, size=10)
        row += 1

    row = _blank(ws, row)
    row = _section(ws, row, "SKIP REASONS (files stage 2 did not run)", 3)
    row = _hdrs(ws, row, ["Skip Reason", "N"])
    for reason, n in s2d["skips"].most_common():
        ws.cell(row, 1, reason)
        _xl(ws.cell(row, 2, n), align="right", num_fmt=_NUM)
        row += 1

    _autofit(ws)


# ----------------------------------------------------------------- sheet: OCR Yield
def _sheet_ocr(wb, rp):
    if not rp.get("ocr_yield"):
        return
    oy = rp["ocr_yield"]
    ws = wb.create_sheet("OCR Yield")
    row = 1
    row = _title(ws, row, "EMBEDDED-IMAGE OCR YIELD", 3)
    row = _blank(ws, row)
    note = ws.cell(row, 1,
                   "Files that gained text only because embedded-image DI was run. "
                   "Suggestive, not causal — re-run with --no-image-ocr and diff lanes "
                   "for the real answer.")
    _xl(note, fg=_C_GTXT, size=10, wrap=True)
    ws.row_dimensions[row].height = 28
    row += 1
    row = _blank(ws, row)

    row = _section(ws, row, "SUMMARY", 3)
    row = _kv(ws, row, "Files that gained embedded-image text", oy["got_count"],
              val_fmt=_NUM)
    row = _kv(ws, row, "DI calls for embedded images", oy["img_calls"], val_fmt=_NUM)
    row = _kv(ws, row, "Share of all DI pages (billable)",
              f"{pct(oy['img_pages'], oy['di_pages_total']):.0%}")
    _xl(ws.cell(row - 1, 3,
                f"embedded-image pages ({oy['img_pages']:,}) ÷ total billable pages "
                f"({oy['di_pages_total']:,}) — DI is billed per page, so this is "
                "the true share of Document Intelligence cost from embedded-image extraction"),
        fg=_C_GTXT, size=10)
    row = _kv(ws, row, "Of those — stage 1 flagged responsive",
              oy["flagged_resp"], val_fmt=_NUM)
    row = _kv(ws, row, "Of those — truly responsive", oy["truly_resp"], val_fmt=_NUM)
    row = _kv(ws, row, "Of those — truly BDE", oy["truly_bde"], val_fmt=_NUM)

    if oy["lanes"]:
        row = _blank(ws, row)
        row = _section(ws, row, "LANE BREAKDOWN", 3)
        row = _hdrs(ws, row, ["Lane", "N"])
        for lane, n in oy["lanes"].most_common():
            ws.cell(row, 1, lane)
            _xl(ws.cell(row, 2, n), align="right", num_fmt=_NUM)
            row += 1

    _autofit(ws)


# ----------------------------------------------------------------- sheet: File Detail
def _sheet_file_detail(wb, rp):
    ws = wb.create_sheet("File Detail")
    COLS = ["Control ID", "File Name", "Type", "Searchable", "Structured",
            "Stage1 Lane", "S1 Call", "Stage2 Lane", "S2 Call", "Pipeline",
            "Truth Resp", "Truth BDE", "Truth Entities",
            "Internal: Rule Entities", "Internal: LLM Persons", "Internal: Effective Entities (BDE tier)",
            "DI Calls", "Tok1", "Tok2", "Elapsed (s)", "Detail"]
    row = _hdrs(ws, 1, COLS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(COLS))}1"

    from openpyxl.styles import PatternFill
    root = rp.get("inventory_root", "")

    def _bool_str(v):
        if v is True:  return "Resp"
        if v is False: return "Non-resp"
        return "Undet"

    for dr in rp["file_detail_rows"]:
        r = row
        fill_color = _CLS_FILL.get(dr["cls"], _C_WHITE)
        fill = PatternFill("solid", fgColor=fill_color)

        vals = [
            dr["cid"], dr["name"], dr["type"],
            "Yes" if dr["searchable"] else "No",
            "Yes" if dr["is_struct"]  else "No",
            dr["lane"], _bool_str(dr["s1_call"]),
            dr["s2_lane"] or "", _bool_str(dr["s2_call"]),
            _bool_str(dr["pipe_call"]),
            "Yes" if dr["truth_resp"] else "No",
            "Yes" if dr["truth_bde"]  else "No",
            dr["human_ent"], dr["tool_est"], dr["bpc"],
            0 if dr["pipe_call"] is False else dr["eff"],
            dr["di"], dr["tok1"], dr["tok2"],
            round(dr["elapsed"], 2), dr["detail"],
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(r, c, val)
            cell.fill = fill
            if isinstance(val, (int, float)) and c > 12:
                cell.alignment = __import__("openpyxl").styles.Alignment(horizontal="right")

        if root and dr.get("rel_path"):
            url = "file://" + os.path.join(root, dr["rel_path"]).replace("\\", "/")
            from openpyxl.styles import Font
            ws.cell(r, 2).hyperlink = url
            ws.cell(r, 2).font = Font(color="0563C1", underline="single")

        row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 44
    for col in "CDEFGHIJKLMNOPQRSTU":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["U"].width = 40  # Detail


# ----------------------------------------------------------------- sheet: Per-file Timing
def _sheet_timing_detail(wb, scaling):
    if scaling is None:
        return
    ws = wb.create_sheet("Per-file Timing")
    m   = scaling["m"]
    win = scaling["win_instances"]
    p_in  = p_out = None   # cost not computed here (no price args available)

    COLS = ["Worker Type", "Worker Instance", "File Name", "Status",
            "Processing (s)", "Attempts", "Tokens IN", "Tokens OUT"]
    row = _hdrs(ws, 1, COLS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H1"

    for i, t in enumerate(sorted(m.tasks, key=lambda x: (x.file_name or "")), 0):
        r = row
        wtype = "windows" if t.worker_instance in win else "linux"
        row_bg = _C_WIN if wtype == "windows" else (_C_ALT if i % 2 == 0 else None)
        vals = [wtype, t.worker_instance or "", t.file_name or "",
                t.status, t.processing_s or "", t.attempt_count,
                t.tokens_in or "", t.tokens_out or ""]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(r, c, val)
            if row_bg:
                cell.fill = __import__("openpyxl").styles.PatternFill(
                    "solid", fgColor=row_bg)
            if isinstance(val, (int, float)) and c >= 5:
                cell.alignment = __import__("openpyxl").styles.Alignment(horizontal="right")
        row += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 16
    for col in "EFGH":
        ws.column_dimensions[col].width = 14


# ----------------------------------------------------------------- sheet: Misses
def _sheet_misses(wb, rp):
    ws = wb.create_sheet("Misses")
    COLS = ["Which", "Control ID", "File Name", "Human Entities", "Type",
            "Stage1 Lane", "Stage2 Lane", "Stage2 Level",
            "Tool Est", "BDE Person Count", "Effective",
            "Searchable", "Structured", "DI Calls", "Detail"]
    row = _hdrs(ws, 1, COLS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(COLS))}1"

    from openpyxl.styles import PatternFill
    _WHICH_COLOR = {
        "stage1_FN":   _C_FN,
        "stage1_FP":   _C_FP,
        "pipeline_FN": _C_FN,
        "bde_FN":      "FFD9B3",
        "bde_FP":      "FFFACD",
    }

    for mr in rp["miss_rows"]:
        r = row
        bg = _WHICH_COLOR.get(mr["which"], _C_WHITE)
        fill = PatternFill("solid", fgColor=bg)
        vals = [mr["which"], mr["cid"], mr["name"], mr["human_ent"],
                mr["type"], mr["lane"], mr["s2_lane"], mr["s2_level"],
                mr["tool_est"], mr["bpc"], mr["eff"],
                "Yes" if mr["searchable"] else "No",
                "Yes" if mr["is_struct"]  else "No",
                mr["di"], mr["detail"]]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(r, c, val)
            cell.fill = fill
        row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 44
    for col in "DEFGHIJKLMN":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["O"].width = 40  # Detail


# ----------------------------------------------------------------- chart layout fix
def _fix_chart_layouts(path):
    """Post-save: inject a ManualLayout into every chart's plotArea.

    openpyxl's chart.plot_area.layout API doesn't serialise to XML, so we
    patch the saved file directly.  The outer-mode layout (no layoutTarget)
    constrains the plotArea width/height so Excel leaves room for axis titles
    and tick labels on all four sides without overlap.
    """
    import zipfile as _zf, io as _io, re as _re

    _LAYOUT = (
        '<layout><manualLayout>'
        '<xMode val="factor"/><yMode val="factor"/>'
        '<x val="0.0"/><y val="0.0"/>'
        '<w val="0.82"/><h val="0.87"/>'
        '</manualLayout></layout>'
    )

    with _zf.ZipFile(path, 'r') as zin:
        names = zin.namelist()
        contents = {n: zin.read(n) for n in names}

    for name in names:
        if not (name.startswith('xl/charts/chart') and name.endswith('.xml')):
            continue
        xml = contents[name].decode('utf-8')
        xml = _re.sub(
            r'(<plotArea>)(<[a-zA-Z])',
            lambda m: m.group(1) + _LAYOUT + m.group(2),
            xml, count=1,
        )
        contents[name] = xml.encode('utf-8')

    buf = _io.BytesIO()
    with _zf.ZipFile(buf, 'w', _zf.ZIP_DEFLATED) as zout:
        for name in names:
            zout.writestr(name, contents[name])
    with open(path, 'wb') as f:
        f.write(buf.getvalue())


# ----------------------------------------------------------------- top-level writer
def write_scorecard_xlsx(rp, scaling, out_path):
    try:
        import openpyxl
    except ImportError:
        sys.stderr.write(
            "warning: openpyxl not installed -- skipping xlsx output.\n"
            "         pip install openpyxl\n")
        return

    wb = openpyxl.Workbook()

    # Sheet order mirrors the plan
    _sheet_run_info(wb, rp)       # sheet 1 (active)
    _sheet_cost(wb, rp, scaling)  # sheet 2
    _sheet_timing(wb, scaling)    # sheet 3
    _sheet_timeline(wb, scaling)  # sheet 4 (charts — skipped if no scaling data)
    _sheet_nrr(wb, rp)            # sheet 5
    _sheet_bde(wb, rp)            # sheet 5
    _sheet_mbt(wb, rp)            # sheet 6
    if rp.get("s2_detail"):
        _sheet_s2(wb, rp)         # sheet 7 (conditional)
    if rp.get("ocr_yield"):
        _sheet_ocr(wb, rp)        # sheet 8 (conditional)
    _sheet_file_detail(wb, rp)    # sheet 9
    _sheet_timing_detail(wb, scaling)  # sheet 10 (skipped if no scaling)
    _sheet_misses(wb, rp)         # sheet 11

    wb.save(out_path)
    _fix_chart_layouts(out_path)
    print(f"\nwrote: {out_path}")


# ======================================================================= main ==
def main():
    ap = argparse.ArgumentParser(
        description="One scorecard for a combined-run inventory: cost, NR/R, BDE.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    ap.add_argument("--inventory", default=DEFAULT_INVENTORY)
    ap.add_argument("--entities", required=True,
                    help="CNG entities export (.csv/.xlsx) with Control ID + Total Entities")
    ap.add_argument("--entities-sheet", default="")
    ap.add_argument("--id-col",    default=DEFAULT_ID_COL)
    ap.add_argument("--count-col", default=DEFAULT_COUNT_COL)
    ap.add_argument("--bde-threshold", type=int, default=DEFAULT_BDE_THRESHOLD)
    ap.add_argument("--absent-means", choices=["auto", "zero", "unreviewed"],
                    default="auto")
    ap.add_argument("--manual", default=None)
    ap.add_argument("--manual-sheet", default="")
    ap.add_argument("--status-col", default=DEFAULT_STATUS_COL)
    ap.add_argument("--price-per-1k-in",    type=float, default=0.0)
    ap.add_argument("--price-per-1k-out",   type=float, default=0.0)
    ap.add_argument("--price-per-1k-pages", type=float, default=0.0,
                    help="Document Intelligence, per 1,000 pages")
    ap.add_argument("--timing", default=None,
                    help="Path to a _timing.json snapshot from collect_outputs.py. "
                         "Auto-detected as <inventory_stem>_timing.json if not provided.")
    ap.add_argument("--out-dir", default=".")
    a = ap.parse_args()

    for p in (a.inventory, a.entities):
        if not os.path.exists(p):
            sys.exit(f"not found: {p}")
    os.makedirs(a.out_dir, exist_ok=True)

    recs, has = load_inventory(a.inventory)
    counts, blanks, dupes = load_entities(a.entities, a.entities_sheet,
                                          a.id_col, a.count_col)

    print("=" * 74)
    print("pii_triage COMBINED-RUN SCORECARD")
    print("=" * 74)
    print(f"inventory : {a.inventory}   ({len(recs):,} files)")
    print(f"entities  : {a.entities}")
    print(f"            {len(counts):,} files with a numeric count"
          + (f", {blanks:,} blank/unparseable (dropped)" if blanks else "")
          + (f", {dupes:,} duplicate Control IDs (kept the larger)" if dupes else ""))
    stage2_present = has["s2_lane"] and any(r["s2_ran"] for r in recs.values())
    print(f"pipeline  : stage 1 {'+ stage 2' if stage2_present else 'ONLY'}")

    zero_rows = sum(1 for v in counts.values() if v == 0)
    if a.absent_means == "auto":
        mode = "unreviewed" if zero_rows else "zero"
        print(f"\nabsent-means: AUTO -> '{mode}'  "
              f"(the export contains {zero_rows:,} zero-entity rows)")
    else:
        mode = a.absent_means
        print(f"\nabsent-means: {mode} (explicit)")

    if mode == "zero":
        scored   = sorted(recs)
        truth_resp = {cid: counts.get(cid, 0) > 0 for cid in scored}
        truth_bde  = {cid: counts.get(cid, 0) > a.bde_threshold for cid in scored}
        human      = {cid: counts.get(cid, 0) for cid in scored}
    else:
        scored   = sorted(cid for cid in recs if cid in counts)
        truth_resp = {cid: counts[cid] > 0 for cid in scored}
        truth_bde  = {cid: counts[cid] > a.bde_threshold for cid in scored}
        human      = {cid: counts[cid] for cid in scored}

    reviewed_not_scanned = sum(1 for cid in counts if cid not in recs)
    overlap    = sum(1 for cid in counts if cid in recs)
    match_rate = pct(overlap, len(counts))
    if len(counts) and match_rate < 0.5:
        print("\n" + "!" * 74)
        print(f"!! ID MATCH RATE IS ONLY {match_rate:.1%}")
        print("!! Check: --id-col, and that Control IDs match file_name minus extension.")
        print("!" * 74)

    print(f"\nscored population   : {len(scored):,}")
    print(f"  truly responsive  : {sum(truth_resp.values()):,}"
          f"  ({pct(sum(truth_resp.values()), len(scored)):.1%})")
    print(f"  truly BDE (> {a.bde_threshold}) : {sum(truth_bde.values()):,}")
    print(f"  in export, not run: {reviewed_not_scanned:,}")
    if mode == "unreviewed":
        print(f"  in run, unreviewed: {len(recs) - len(scored):,}")

    # =============================== 1. COST =================================
    di_calls = sum(r["di"]        for r in recs.values())
    di_pages = sum(r["ocr_pages"] + r["img_c"] for r in recs.values())
    ff_calls = sum(1 for r in recs.values() if r["ocr_att"])
    img_calls = sum(r["img_c"]   for r in recs.values())
    img_q     = sum(r["img_q"]   for r in recs.values())
    img_ok    = sum(r["img_ok"]  for r in recs.values())
    img_fail  = sum(r["img_fail"]for r in recs.values())
    tok1      = sum(r["tok1"]    for r in recs.values())
    tok2      = sum(r["tok2"]    for r in recs.values())
    n_llm1    = sum(1 for r in recs.values() if r["llm_consulted"])
    n_llm2    = sum(1 for r in recs.values() if r["s2_ran"] and r["tok2"] > 0)

    print("\n" + "=" * 74)
    print("1. RUN COST")
    print("=" * 74)
    print(f"  DI calls {di_calls:,}   billable pages {di_pages:,}")
    print(f"  stage-1 tokens {tok1:,}   stage-2 tokens {tok2:,}   "
          f"total {tok1 + tok2:,}")

    di_cost = (di_pages / 1000.0 * a.price_per_1k_pages
               if a.price_per_1k_pages else None)
    lo, hi  = sorted((a.price_per_1k_in, a.price_per_1k_out))
    llm_lo  = (tok1 + tok2) / 1000.0 * lo if hi else None
    llm_hi  = (tok1 + tok2) / 1000.0 * hi if hi else None

    if di_cost is not None:
        print(f"  DI cost ${di_cost:,.2f}")
    if llm_hi is not None:
        print(f"  LLM cost ${llm_lo:,.4f} - ${llm_hi:,.4f}")
    if di_cost is not None and llm_hi is not None:
        print(f"  TOTAL (upper) ${di_cost + llm_hi:,.2f}")

    surv      = [cid for cid in recs if not recs[cid]["nr1"]]
    dup_di    = sum(recs[c]["di"] for c in surv)
    dup_pages = sum(recs[c]["ocr_pages"] + recs[c]["img_c"] for c in surv)
    print(f"\n  single-pass saving: {len(surv):,} survivors of {len(recs):,}"
          f"  ({pct(len(surv), len(recs)):.1%})")
    print(f"    DI calls 2nd pass would repeat: {dup_di:,}  ({dup_pages:,} pages)")
    print(f"    saved {pct(dup_di, di_calls + dup_di):.1%} of total DI calls")

    # ============================= 2. NR / R =================================
    s1 = stats(confusion(scored, recs, truth_resp, stage1_call))
    print_rnr("STAGE 1 (Anna's) -- all scored files", s1,
              "population: all scored files")
    nr_results = {"Stage 1 (all scored)": s1}
    nr_pops    = {"Stage 1 (all scored)":
                  "all scored files — the number your tracker uses"}
    s2_pop = []
    if stage2_present:
        s2_pop = [cid for cid in scored if recs[cid]["s2_ran"]]
        s2 = stats(confusion(s2_pop, recs, truth_resp, stage2_call))
        print_rnr("STAGE 2 (Daniel's) -- graded files only", s2,
                  f"population: {len(s2_pop):,} files stage 2 graded")
        nr_results["Stage 2 (graded)"] = s2
        _s2_undet_note = (f"; {s2['undetermined']:,} undetermined → N = {s2['N']:,}"
                          if s2["undetermined"] else "")
        nr_pops["Stage 2 (graded)"] = (
            f"{len(s2_pop):,} files stage 2 actually graded{_s2_undet_note}"
        )

        s1s = stats(confusion(s2_pop, recs, truth_resp, stage1_call))
        print_rnr("STAGE 1 (restricted to stage 2's population)", s1s,
                  "apples-to-apples with the row above")
        nr_results["Stage 1 (S2 population)"] = s1s
        # Stage 1 may be undetermined for files in non-standard lanes (e.g.
        # nonsearchable_sample, review_error) that Stage 2 still graded — those
        # files are excluded from N, which is why N here can be < Stage 2 N.
        _s1s_undet_note = (f"; {s1s['undetermined']:,} have non-standard S1 lanes "
                           f"(undetermined) → N = {s1s['N']:,}"
                           if s1s["undetermined"] else "")
        nr_pops["Stage 1 (S2 population)"] = (
            f"same {len(s2_pop):,} files stage 2 graded{_s1s_undet_note} "
            f"— compare with Stage 2 row above"
        )

        pipe = stats(confusion(scored, recs, truth_resp, pipeline_call))
        print_rnr("PIPELINE (sequential — what a reviewer receives)", pipe)
        nr_results["Pipeline"] = pipe
        nr_pops["Pipeline"]    = \
            "all scored: S1 clears→gone; else S2's call; if S2 N/A, S1 stands"

        uni = stats(confusion(scored, recs, truth_resp, union_call))
        print_rnr("UNION (either stage says responsive)", uni)
        nr_results["Union"] = uni
        nr_pops["Union"]    = "recall ceiling — not the workflow, shows gating cost"

        print("\n  stage-2 level distribution:")
        for lvl, n in Counter(recs[c]["s2_level"] or "(none)"
                               for c in s2_pop).most_common():
            tr = sum(1 for c in s2_pop
                     if recs[c]["s2_level"] == lvl and truth_resp[c])
            print(f"    {lvl:12s} {n:7,}   truly resp {tr:,} ({pct(tr, n):.1%})")
        print("\n  stage-2 skip reasons:")
        for k, n in Counter(recs[c]["s2_skip"] or "(none)"
                             for c in scored if not recs[c]["s2_ran"]).most_common():
            print(f"    {k:20s} {n:7,}")

    # optional manual cross-check
    xcheck = {}
    if a.manual:
        man, unrec = load_manual_status(a.manual, a.manual_sheet,
                                        a.id_col, a.status_col)
        common = [cid for cid in scored if cid in man]
        agree  = sum(1 for cid in common
                     if (man[cid] == "resp") == truth_resp[cid])
        print(f"\n  truth cross-check: {len(common):,} common files, "
              f"{agree:,} agree ({pct(agree, len(common)):.1%})")
        xcheck = {
            "xcheck_n": len(common),
            "xcheck_agree_pct": f"{pct(agree, len(common)):.1%}",
            "xcheck_disagree": Counter(
                ("entities>0, reviewer=No" if truth_resp[c] else
                 "entities=0, reviewer=Yes")
                for c in common if (man[c] == "resp") != truth_resp[c]
            ).most_common() if agree < len(common) else [],
        }

    # =============================== 3. BDE =================================
    print("\n" + "=" * 74)
    print(f"3. BDE ACCURACY   (truth: Total Entities > {a.bde_threshold})")
    print("=" * 74)
    bde_defs = {
        "is_bde flag (column)":
            lambda x: x["is_bde"],
        "BDE lane (bde / structured_bde)":
            lambda x: x["lane"] in BDE_LANES,
        f"effective count > {a.bde_threshold}":
            lambda x: x["eff"] > a.bde_threshold,
        f"lane OR bde_person_count > {a.bde_threshold} (incl. recovery)":
            lambda x: x["lane"] in BDE_LANES or x["bpc"] > a.bde_threshold,
    }
    if stage2_present:
        bde_defs["stage 2 s2_is_bde"] = lambda x: x["s2_is_bde"]

    bde_results = {}
    for label, pred in bde_defs.items():
        c = confusion(scored, recs, truth_bde, lambda x, p=pred: p(x))
        st = stats(c)
        bde_results[label] = (st, c)
        print_bde(label, st)

    # BDE count accuracy
    true_bde_ids = [cid for cid in scored if truth_bde[cid]]
    count_acc = None
    if true_bde_ids:
        thr1   = a.bde_threshold + 1
        raw_ge = sum(1 for c in true_bde_ids if recs[c]["est"] >= thr1)
        eff_ge = sum(1 for c in true_bde_ids if recs[c]["eff"] >= thr1)
        err    = sorted(abs(recs[c]["eff"] - human[c]) for c in true_bde_ids)
        under  = [c for c in true_bde_ids if recs[c]["eff"] < thr1]
        rescued = eff_ge - raw_ge
        print(f"\n  COUNT ACCURACY on {len(true_bde_ids):,} true BDEs")
        print(f"    raw est >= {thr1}: {raw_ge:,}  ({pct(raw_ge, len(true_bde_ids)):.1%})")
        print(f"    effective >= {thr1}: {eff_ge:,}  ({pct(eff_ge, len(true_bde_ids)):.1%})")
        print(f"    rescued by bde_person_count: {rescued:,}")
        print(f"    still missed: {len(under):,}  ({pct(len(under), len(true_bde_ids)):.1%})")
        print(f"    mean |tool-human|: {sum(err)/len(err):,.1f}"
              f"    median: {err[len(err)//2]:,}")
        count_acc = {
            "n_true_bde":       len(true_bde_ids),
            "threshold":        thr1,
            "raw_ge":           raw_ge,
            "eff_ge":           eff_ge,
            "rescued":          rescued,
            "still_missed":     len(under),
            "mean_err":         sum(err) / len(err),
            "median_err":       err[len(err) // 2],
            "under_by_type":    Counter(recs[c]["type"] for c in under),
            "under_by_searchable": Counter(
                "searchable" if recs[c]["searchable"] else "non-searchable"
                for c in under),
        }

    # =============================== 4. BREAKDOWNS ==========================
    print("\n" + "=" * 74)
    print("4. BY FILE TYPE (stage 1 R/NR)")
    print("=" * 74)
    hdr = (f"{'TYPE':<14}{'n':>7}{'TP':>7}{'FP':>7}{'FN':>7}{'TN':>7}"
           f"{'Rec':>8}{'Prec':>8}{'NRacc':>8}{'DIcalls':>9}{'tokens':>12}")
    print(hdr)
    print("-" * len(hdr))
    by = defaultdict(list)
    for cid in scored:
        by[recs[cid]["type"]].append(cid)
    for t, ids in sorted(by.items(), key=lambda kv: -len(kv[1])):
        st = stats(confusion(ids, recs, truth_resp, stage1_call))
        print(f"{t:<14}{st['N']:>7,}{st['TP']:>7,}{st['FP']:>7,}"
              f"{st['FN']:>7,}{st['TN']:>7,}"
              f"{fmt(st['recall'],3):>8}{fmt(st['precision'],3):>8}"
              f"{fmt(st['nr_accuracy'],3):>8}"
              f"{sum(recs[c]['di'] for c in ids):>9,}"
              f"{sum(recs[c]['tok1']+recs[c]['tok2'] for c in ids):>12,}")

    # OCR yield
    ocr_yield = None
    if has["img_ocr_calls"]:
        got = [cid for cid in scored if recs[cid]["img_ok"] > 0]
        print(f"\n  OCR YIELD: {len(got):,} files gained embedded-image text"
              f"  ({img_calls:,} DI calls, {pct(img_calls, max(di_calls, 1)):.0%} of DI spend)")
        if got:
            ocr_yield = {
                "got_count":      len(got),
                "img_calls":      img_calls,
                "img_pages":      img_calls,   # 1 embedded-image DI call = 1 page
                "di_pages_total": di_pages,
                "di_calls_total": di_calls,
                "flagged_resp":   sum(1 for c in got if not recs[c]["nr1"]),
                "truly_resp":     sum(1 for c in got if truth_resp[c]),
                "truly_bde":      sum(1 for c in got if truth_bde[c]),
                "lanes":          Counter(recs[c]["lane"] for c in got),
            }

    # ==================== collect per-file breakdown data ===================
    def _new_mbt():
        return {
            "ftype":   defaultdict(lambda: [0, 0, 0, 0]),
            "search":  defaultdict(lambda: [0, 0, 0, 0]),
            "sftype":  defaultdict(lambda: defaultdict(lambda: [0, 0, 0, 0])),
            "sp":      defaultdict(lambda: [0, 0, 0, 0]),
            "spftype": defaultdict(lambda: defaultdict(lambda: [0, 0, 0, 0])),
            "overall": [0, 0, 0, 0],
        }
    mbt = {"s1": _new_mbt(), "s2": _new_mbt(), "pipe": _new_mbt()}

    file_detail_rows = []

    for cid in scored:
        rec = recs[cid]
        s1  = stage1_call(rec)
        tr  = truth_resp[cid]
        tb  = truth_bde[cid]
        ftype = rec["type"]
        skey  = "searchable" if rec["searchable"] else "non-searchable"
        pkey  = "structured" if rec["programmatic"] else "non-structured"

        pc = pipeline_call(rec) if stage2_present else s1

        calls_to_record = [("s1", s1)]
        if stage2_present:
            calls_to_record.append(("s2", stage2_call(rec)))
            calls_to_record.append(("pipe", pc))

        for bk, call in calls_to_record:
            if call is None:
                continue
            if   tr and call:       idx = 0
            elif call and not tr:   idx = 1
            elif tr and not call:   idx = 2
            else:                   idx = 3
            b = mbt[bk]
            b["overall"][idx]                     += 1
            b["ftype"][ftype][idx]                += 1
            b["search"][skey][idx]                += 1
            b["sftype"][skey][ftype][idx]         += 1
            b["sp"][(skey, pkey)][idx]            += 1
            b["spftype"][(skey, pkey)][ftype][idx]+= 1
        if pc is None:
            pcls = "Undetermined"
        elif tr and pc:     pcls = "TP"
        elif pc and not tr: pcls = "FP"
        elif tr and not pc: pcls = "FN"
        else:               pcls = "TN"

        file_detail_rows.append({
            "cid":        cid,
            "name":       rec["name"],
            "rel_path":   rec["rel_path"],
            "type":       rec["type"],
            "searchable": rec["searchable"],
            "is_struct":  rec["is_struct"],
            "lane":       rec["lane"],
            "s1_call":    s1,
            "s2_lane":    rec["s2_lane"],
            "s2_call":    stage2_call(rec),
            "pipe_call":  pc,
            "cls":        pcls,
            "truth_resp": tr,
            "truth_bde":  tb,
            "human_ent":  human.get(cid, ""),
            "tool_est":   rec["est"],
            "bpc":        rec["bpc"],
            "eff":        rec["eff"],
            "di":         rec["di"],
            "tok1":       rec["tok1"],
            "tok2":       rec["tok2"],
            "elapsed":    rec["elapsed"],
            "detail":     rec["detail"],
            "s2_level":   rec["s2_level"],
        })

    # zero-fill the ftype buckets for all 9 known types so every block has every row
    for bk in ("s1", "s2", "pipe"):
        b = mbt[bk]
        for ft in _FTYPE_ORDER:
            _ = b["ftype"][ft]
            for sk in ("searchable", "non-searchable"):
                _ = b["sftype"][sk][ft]
                for pk in ("structured", "non-structured"):
                    _ = b["spftype"][(sk, pk)][ft]

    # Stage 2 detail
    s2_detail = None
    if stage2_present and s2_pop:
        levels = []
        for lvl, n in Counter(recs[c]["s2_level"] or "(none)"
                               for c in s2_pop).most_common():
            tr_c = sum(1 for c in s2_pop
                       if recs[c]["s2_level"] == lvl and truth_resp[c])
            levels.append({"level": lvl, "n": n, "true_resp": tr_c,
                            "true_resp_pct": pct(tr_c, n)})
        s2_detail = {
            "levels": levels,
            "skips":  Counter(recs[c]["s2_skip"] or "(none)"
                               for c in scored if not recs[c]["s2_ran"]),
        }

    # Miss rows
    c1 = confusion(scored, recs, truth_resp, stage1_call)
    bde_key = f"lane OR bde_person_count > {a.bde_threshold} (incl. recovery)"
    _, cb = bde_results[bde_key]

    miss_groups = [("stage1_FN", c1["fns"]), ("stage1_FP", c1["fps"])]
    if stage2_present:
        cp = confusion(scored, recs, truth_resp, pipeline_call)
        miss_groups.append(("pipeline_FN", cp["fns"]))
    miss_groups += [("bde_FN", cb["fns"]), ("bde_FP", cb["fps"])]

    miss_rows = []
    for which, cids in miss_groups:
        for cid in sorted(cids, key=lambda c: -human.get(c, 0)):
            rec = recs[cid]
            miss_rows.append({
                "which":     which,    "cid":       cid,
                "name":      rec["name"], "human_ent": human.get(cid, ""),
                "type":      rec["type"], "lane":      rec["lane"],
                "s2_lane":   rec["s2_lane"], "s2_level":  rec["s2_level"],
                "tool_est":  rec["est"], "bpc":       rec["bpc"],
                "eff":       rec["eff"],
                "searchable": rec["searchable"], "is_struct": rec["is_struct"],
                "di":        rec["di"], "detail":     rec["detail"],
            })

    # ==================== try scaling data ==================================
    _auto_timing = os.path.splitext(a.inventory)[0] + "_timing.json"
    timing_path = a.timing or (_auto_timing if os.path.exists(_auto_timing) else None)
    if timing_path:
        print(f"\nLoading timing snapshot: {timing_path}...", end=" ", flush=True)
        try:
            scaling = _load_timing_snapshot(timing_path)
            print("ok")
        except Exception as _exc:
            print(f"failed ({_exc}), falling back to live run_metrics()...")
            scaling = _try_run_metrics()
            print("ok" if scaling else "not available")
    else:
        print("\nFetching scaling-lib run metrics...", end=" ", flush=True)
        scaling = _try_run_metrics()
        print("ok" if scaling else "not available")

    # ==================== write Excel =======================================
    today = datetime.date.today().strftime("%Y%m%d")
    out_path = os.path.join(a.out_dir, f"scorecard_{today}.xlsx")

    report = {
        # run info
        "inventory_path":    a.inventory,
        "entities_path":     a.entities,
        "manual_path":       a.manual,
        "generated":         datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "pipeline_desc":     ("stage 1 + stage 2" if stage2_present
                               else "stage 1 only"),
        "mode":              mode,
        "absent_auto":       a.absent_means == "auto",
        "bde_threshold":     a.bde_threshold,
        # population
        "n_inventory":       len(recs),
        "n_scored":          len(scored),
        "n_responsive":      sum(truth_resp.values()),
        "n_bde":             sum(truth_bde.values()),
        "reviewed_not_scanned": reviewed_not_scanned,
        "n_not_reviewed":    len(recs) - len(scored) if mode == "unreviewed" else 0,
        "match_rate":        match_rate,
        # cost
        "di_calls":          di_calls,
        "di_pages":          di_pages,
        "img_calls":         img_calls,
        "img_q":             img_q,
        "img_ok":            img_ok,
        "img_fail":          img_fail,
        "tok1":              tok1,
        "tok2":              tok2,
        "n_llm1":            n_llm1,
        "n_llm2":            n_llm2,
        "di_cost":           di_cost,
        "llm_lo":            llm_lo,
        "llm_hi":            llm_hi,
        "price_per_1k_in":   a.price_per_1k_in,
        "price_per_1k_out":  a.price_per_1k_out,
        "price_per_1k_pages": a.price_per_1k_pages,
        "n_surv":            len(surv),
        "dup_di":            dup_di,
        "dup_pages":         dup_pages,
        # NR/R
        "nr_results":        nr_results,
        "nr_pops":           nr_pops,
        # BDE
        "bde_results":       bde_results,
        "count_acc":         count_acc,
        # breakdowns — s1/s2/pipe keys consumed by _sheet_mbt
        "stage2_present":    stage2_present,
        "mbt_s1_overall":    mbt["s1"]["overall"],
        "mbt_s1_ftype":      mbt["s1"]["ftype"],
        "mbt_s1_search":     mbt["s1"]["search"],
        "mbt_s1_sftype":     mbt["s1"]["sftype"],
        "mbt_s1_sp":         mbt["s1"]["sp"],
        "mbt_s1_spftype":    mbt["s1"]["spftype"],
        "mbt_s2_overall":    mbt["s2"]["overall"],
        "mbt_s2_ftype":      mbt["s2"]["ftype"],
        "mbt_s2_search":     mbt["s2"]["search"],
        "mbt_s2_sftype":     mbt["s2"]["sftype"],
        "mbt_s2_sp":         mbt["s2"]["sp"],
        "mbt_s2_spftype":    mbt["s2"]["spftype"],
        "mbt_pipe_overall":  mbt["pipe"]["overall"],
        "mbt_pipe_ftype":    mbt["pipe"]["ftype"],
        "mbt_pipe_search":   mbt["pipe"]["search"],
        "mbt_pipe_sftype":   mbt["pipe"]["sftype"],
        "mbt_pipe_sp":       mbt["pipe"]["sp"],
        "mbt_pipe_spftype":  mbt["pipe"]["spftype"],
        # stage 2
        "s2_detail":         s2_detail,
        # OCR yield
        "ocr_yield":         ocr_yield,
        # per-file
        "file_detail_rows":  file_detail_rows,
        "miss_rows":         miss_rows,
        # misc
        "inventory_root":    _read_manifest_root(a.inventory),
        "_recs":             recs,   # used by Cost sheet for some per-file sums
        "worker_cost":       _compute_worker_cost(
            scaling["completed"], scaling["win_instances"],
            scaling.get("worker_config"),
        ) if scaling else None,
        **xcheck,
    }

    write_scorecard_xlsx(report, scaling, out_path)

    print("\nREAD THE MISSES FIRST. A stage-1 FN is a document with real PII that the tool")
    print("cleared -- operationally a missed notification, and the only error class that")
    print("cannot be caught downstream.")


if __name__ == "__main__":
    main()
